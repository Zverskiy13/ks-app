# -*- coding: utf-8 -*-
"""
«Клиники Столицы» — API + хостинг PWA (один сервис на Railway).
Читает реальное состояние бота из GitHub-репозитория и отдаёт его фронту
с ролевой фильтрацией. Также раздаёт сам фронт (папка ./webapp).

ENV (Railway → Variables):
  GITHUB_TOKEN   — тот же fine-grained PAT, что у бота (Contents: read)
  GITHUB_REPO    — например Zverskiy13/ks-telegram-bot
  GITHUB_BRANCH  — main
  APP_PINS       — ОБЯЗАТЕЛЬНО (или APP_PIN_HASHES): JSON {pin: {id,name,role,companies,title}}.
                   Без него приложение НЕ стартует — демо-PIN в коде отсутствуют.
  APP_PIN_HASHES — альтернатива APP_PINS: JSON [{id,...,salt(hex),hash(hex)}] (без plaintext-PIN).
  SECRET_KEY     — ключ подписи сессии (опц.; иначе стабильно из GITHUB_TOKEN).
Старт: uvicorn app:app --host 0.0.0.0 --port $PORT   (см. Procfile)
"""
import os, json, base64, re, threading, time as _time, datetime as dt
import hmac, hashlib, secrets
from fastapi import FastAPI, HTTPException, Request, Response, Depends
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import requests
try:
    import db as _db                    # Postgres-слой (Фаза 0). Спит, если DATABASE_URL не задан.
except Exception:
    _db = None

GH_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GH_REPO = os.environ.get("GITHUB_REPO", "")
GH_BRANCH = os.environ.get("GITHUB_BRANCH", "main")

# ---------- Web Push (VAPID) ----------
VAPID_PUBLIC = os.environ.get("VAPID_PUBLIC", "BM9xFCyzg-emFABRWFgNt3-1ChYdUhhCgfZC5FZAumBhvidqJ3eeLNRdcMC8Sx-2sWh91PjM3NwAywvJGh00PT8")
VAPID_SUB = os.environ.get("VAPID_SUB", "mailto:admin@kliniki-stolicy.ru")
_VAPID_PEM = "/tmp/vapid.pem"


def _normalize_pem(raw):
    """Принять ключ как угодно: многострочный PEM, PEM с \\n, или просто base64 одной строкой."""
    raw = (raw or "").strip().replace("\\n", "\n")
    m = re.search(r"-----BEGIN ([A-Z ]+)-----(.*?)-----END \1-----", raw, re.S)
    if m:
        label, body = m.group(1).strip(), re.sub(r"\s+", "", m.group(2))
    else:
        label, body = "PRIVATE KEY", re.sub(r"\s+", "", raw)
    wrapped = "\n".join(body[i:i + 64] for i in range(0, len(body), 64))
    return f"-----BEGIN {label}-----\n{wrapped}\n-----END {label}-----\n"


if os.environ.get("VAPID_PRIVATE"):
    try:
        with open(_VAPID_PEM, "w") as _f:
            _f.write(_normalize_pem(os.environ["VAPID_PRIVATE"]))
    except Exception:
        pass
PUSH_HOUR_UTC = int(os.environ.get("PUSH_HOUR_UTC", "5"))   # ≈8:00 МСК — утренний пуш


def push_ready():
    return os.path.exists(_VAPID_PEM)


def send_push(sub, payload):
    from pywebpush import webpush
    webpush(subscription_info=sub, data=json.dumps(payload, ensure_ascii=False),
            vapid_private_key=_VAPID_PEM, vapid_claims={"sub": VAPID_SUB})

ROLE_SECTIONS = {
    "owner": ["home", "day", "tasks", "journal", "money", "funnel", "more"],
    "head":  ["home", "day", "tasks", "journal", "money", "funnel", "more"],
    "staff": ["home", "day", "tasks", "journal", "more"],
}
# Пользователи и PIN берутся ТОЛЬКО из окружения. Демо-значений в коде нет.
# APP_PINS       — JSON {pin: {id,name,role,companies,title}} (plaintext-PIN только в env)
# APP_PIN_HASHES — JSON [{id,name,role,companies,title, salt(hex), hash(hex)}] (без plaintext)
_APP_PINS_RAW = os.environ.get("APP_PINS")
_APP_PIN_HASHES_RAW = os.environ.get("APP_PIN_HASHES")
if not (_APP_PINS_RAW or _APP_PIN_HASHES_RAW):
    raise RuntimeError("Запуск запрещён: не заданы APP_PINS или APP_PIN_HASHES. "
                       "В коде нет демо-PIN — задайте пользователей через переменные окружения.")

app = FastAPI(title="Клиники Столицы")
_ORIGINS = [o.strip() for o in os.environ.get("APP_ORIGIN", "").split(",") if o.strip()]
app.add_middleware(CORSMiddleware, allow_origins=_ORIGINS, allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])


# ================= БЕЗОПАСНОСТЬ: сессии (HttpOnly cookie) + scrypt для PIN =================
SECRET_KEY = os.environ.get("SECRET_KEY") or hashlib.sha256(("ks-sess-" + (GH_TOKEN or "local-dev")).encode()).hexdigest()
# ротация ключа: подписываем текущим, проверяем текущим + старым (SECRET_KEY_OLD)
_SECRETS = [SECRET_KEY.encode()] + ([os.environ["SECRET_KEY_OLD"].encode()] if os.environ.get("SECRET_KEY_OLD") else [])
SESSION_COOKIE = "ks_session"
SESSION_TTL = int(os.environ.get("SESSION_TTL_DAYS", "30")) * 86400
SESSION_SECURE = os.environ.get("SESSION_INSECURE", "") != "1"   # по умолчанию Secure (только HTTPS)


def _scrypt(pin, salt):
    return hashlib.scrypt(str(pin).encode("utf-8"), salt=salt, n=16384, r=8, p=1, dklen=32)


def _b64u(b):
    return base64.urlsafe_b64encode(b).decode().rstrip("=")


def _b64u_dec(s):
    return base64.urlsafe_b64decode(s + "=" * (-len(s) % 4))


def _build_pin_table():
    """Таблица scrypt-хешей PIN. Источник — ТОЛЬКО окружение (APP_PIN_HASHES или APP_PINS)."""
    table = []
    if _APP_PIN_HASHES_RAW:
        for rec in json.loads(_APP_PIN_HASHES_RAW):
            table.append({"salt": bytes.fromhex(rec["salt"]), "hash": bytes.fromhex(rec["hash"]),
                          "profile": {k: rec[k] for k in ("id", "name", "role", "companies", "title") if k in rec}})
        return table
    for pin, prof in json.loads(_APP_PINS_RAW).items():
        salt = secrets.token_bytes(16)
        table.append({"salt": salt, "hash": _scrypt(pin, salt), "profile": prof})
    return table


PIN_TABLE = _build_pin_table()
USERS_BY_ID = {rec["profile"]["id"]: rec["profile"] for rec in PIN_TABLE}


def verify_pin(pin):
    for rec in PIN_TABLE:
        if hmac.compare_digest(_scrypt(pin or "", rec["salt"]), rec["hash"]):
            return rec["profile"]
    return None


def make_session(uid):
    exp = int(_time.time()) + SESSION_TTL
    payload = _b64u(json.dumps({"uid": uid, "exp": exp}, separators=(",", ":")).encode())
    sig = _b64u(hmac.new(_SECRETS[0], payload.encode(), hashlib.sha256).digest())
    return payload + "." + sig


def read_session(token):
    if not token or "." not in token:
        return None
    payload, sig = token.rsplit(".", 1)
    if not any(hmac.compare_digest(sig, _b64u(hmac.new(sec, payload.encode(), hashlib.sha256).digest())) for sec in _SECRETS):
        return None
    try:
        data = json.loads(_b64u_dec(payload))
    except Exception:
        return None
    if int(data.get("exp", 0)) < int(_time.time()):
        return None
    return data


def _set_session_cookie(resp, uid):
    resp.set_cookie(SESSION_COOKIE, make_session(uid), max_age=SESSION_TTL,
                    httponly=True, secure=SESSION_SECURE, samesite="lax", path="/")


def _clear_session_cookie(resp):
    resp.delete_cookie(SESSION_COOKIE, path="/")


# антибрутфорс входа (в памяти, по IP)
_LOGIN_FAILS = {}


def _login_gate(ip):
    rec = _LOGIN_FAILS.get(ip)
    return not (rec and rec.get("until", 0) > _time.time())


def _login_fail(ip):
    rec = _LOGIN_FAILS.get(ip, {"n": 0, "until": 0})
    rec["n"] += 1
    if rec["n"] >= 5:
        rec["until"] = _time.time() + 300   # блок на 5 минут после 5 неудач
        rec["n"] = 0
    _LOGIN_FAILS[ip] = rec


def _login_ok(ip):
    _LOGIN_FAILS.pop(ip, None)


def _token_from(request):
    tok = request.cookies.get(SESSION_COOKIE)
    if not tok:
        auth = request.headers.get("authorization", "")
        if auth.startswith("Bearer "):
            tok = auth[7:]
    return tok


def current_user(request: Request):
    data = read_session(_token_from(request))
    if not data:
        raise HTTPException(401, "Не авторизовано")
    return by_id(data["uid"])


def require_owner(request: Request):
    u = current_user(request)
    if u.get("role") != "owner":
        raise HTTPException(403, "Доступно только владельцу")
    return u


def _prof(u):
    return u if isinstance(u, dict) else by_id(u)


_AUTH_WHITELIST = {"/api/login", "/api/auth/login", "/api/auth/me", "/api/auth/logout"}


@app.middleware("http")
async def _auth_guard(request: Request, call_next):
    p = request.url.path
    if p.startswith("/api/") and p not in _AUTH_WHITELIST:
        if not read_session(_token_from(request)):
            return JSONResponse({"ok": False, "error": "Не авторизовано"}, status_code=401)
    return await call_next(request)


MAX_BODY = int(os.environ.get("MAX_BODY_MB", "15")) * 1024 * 1024
_CSP = ("default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://unpkg.com https://cdn.jsdelivr.net; "
        "font-src 'self' https://fonts.gstatic.com https://unpkg.com https://cdn.jsdelivr.net; "
        "img-src 'self' data: blob:; "
        "frame-src 'self' data: blob:; "
        "connect-src 'self'; object-src 'none'; base-uri 'self'; "
        "form-action 'self'; frame-ancestors 'none'")


@app.middleware("http")
async def _security_headers(request: Request, call_next):
    if request.method in ("POST", "PUT", "PATCH"):
        cl = request.headers.get("content-length", "")
        if cl.isdigit() and int(cl) > MAX_BODY:
            return JSONResponse({"ok": False, "error": "Слишком большой запрос"}, status_code=413)
    resp = await call_next(request)
    resp.headers["Content-Security-Policy"] = _CSP
    resp.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["X-Frame-Options"] = "DENY"
    resp.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    resp.headers["Permissions-Policy"] = "geolocation=(), camera=(self), microphone=(self)"
    return resp


_GH_CACHE = {}        # path -> (text, ts)
_GH_TTL = 25          # сек: чтения из памяти, без обращения к GitHub


def _storage_db(path):
    """True — этот путь хранить в Postgres (флаг STORAGE_BACKEND=db).
       Данные бота (ПВЛ, state/workload/*) всегда остаются на GitHub: их пишет бот,
       приложению они нужны «живыми», а не разово перенесёнными."""
    if not (_db and _db.db_available()):
        return False
    if os.environ.get("STORAGE_BACKEND", "github") != "db":
        return False
    # Общие с ботом файлы (пишет бот). Пока бот на GitHub (BOT_ON_GITHUB=1, по умолч.) —
    # читаем их из GitHub «живьём». Когда бот переедет на общий Postgres — выставить
    # BOT_ON_GITHUB=0, и приложение будет брать их из общего KV.
    if os.environ.get("BOT_ON_GITHUB", "1") == "1" and (
            path.startswith("state/workload/") or path == "state/heartbeat.json"):
        return False
    return True


def _gh_read_github(path, fresh=False):
    now = _time.time()
    if not fresh:
        c = _GH_CACHE.get(path)
        if c and (now - c[1]) < _GH_TTL:
            return c[0]
    if not (GH_TOKEN and GH_REPO):
        return None
    try:
        r = requests.get(f"https://api.github.com/repos/{GH_REPO}/contents/{path}?ref={GH_BRANCH}",
                         headers={"Authorization": f"token {GH_TOKEN}"}, timeout=15)
        if r.status_code == 200:
            txt = base64.b64decode(r.json()["content"]).decode("utf-8")
            _GH_CACHE[path] = (txt, now)
            return txt
    except Exception:
        pass
    return None


def gh_read(path, fresh=False):
    if _storage_db(path):
        try:
            v = _db.kv_get(path)
        except Exception:
            v = None
        if v is not None:
            return v
        # ленивая миграция: в KV пусто — тянем из GitHub и переносим в Postgres
        gv = _gh_read_github(path, fresh=True)
        if gv is not None:
            try:
                _db.kv_set(path, gv)
            except Exception:
                pass
        return gv
    return _gh_read_github(path, fresh)


def load_json(path, default):
    txt = gh_read(path)
    try:
        return json.loads(txt) if txt else default
    except Exception:
        return default


def gh_stat_read(path):
    """(text, status): 200 — файл есть; 404 — файла нет; иначе (5xx/сеть/лимит) — чтение не удалось."""
    if _storage_db(path):
        try:
            v = _db.kv_get(path)
        except Exception:
            return None, -1                  # БД недоступна — чтение не удалось, писать нельзя
        if v is not None:
            return v, 200
        # в KV нет — попробуем перенести из GitHub (первый доступ), иначе считаем «файла нет»
        gv = _gh_read_github(path, fresh=True)
        if gv is not None:
            try:
                _db.kv_set(path, gv)
            except Exception:
                pass
            return gv, 200
        return None, 404
    if not (GH_TOKEN and GH_REPO):
        return None, 0
    try:
        r = requests.get(f"https://api.github.com/repos/{GH_REPO}/contents/{path}?ref={GH_BRANCH}",
                         headers={"Authorization": f"token {GH_TOKEN}"}, timeout=20)
        if r.status_code == 200:
            return base64.b64decode(r.json()["content"]).decode("utf-8"), 200
        return None, r.status_code
    except Exception:
        return None, -1


def load_store_safe(path):
    """Безопасное чтение JSON-словаря перед перезаписью (read-modify-write).
       Возвращает (store, safe). safe=False — чтение сорвалось, писать НЕЛЬЗЯ
       (иначе затрём уже занесённые данные пустым словарём)."""
    if not _storage_db(path) and not (GH_TOKEN and GH_REPO):
        return {}, True                  # хранилище отключено — прежнее поведение
    for _ in range(3):
        txt, st = gh_stat_read(path)
        if st == 200:
            try:
                data = json.loads(txt)
                return (data if isinstance(data, dict) else {}), True
            except Exception:
                return {}, True          # файл есть, но битый JSON — не блокируем занос
        if st == 404:
            return {}, True              # файла ещё нет (новый месяц) — создавать можно
        _time.sleep(1.5)                 # 5xx / сеть / лимит API — подождать и повторить
    return {}, False                     # прочитать так и не смогли — запись запрещаем


def gh_delete(path, message):
    """Удалить файл из репозитория-хранилища."""
    if _storage_db(path):
        try:
            _db.kv_del(path)
            _GH_CACHE.pop(path, None)
            return True
        except Exception:
            return False
    if not (GH_TOKEN and GH_REPO):
        return False
    url = f"https://api.github.com/repos/{GH_REPO}/contents/{path}"
    h = {"Authorization": f"token {GH_TOKEN}", "Accept": "application/vnd.github+json"}
    try:
        r = requests.get(f"{url}?ref={GH_BRANCH}", headers=h, timeout=15)
        if r.status_code != 200:
            return True  # файла и так нет
        sha = r.json().get("sha")
        pr = requests.delete(url, headers=h,
                             json={"message": message, "branch": GH_BRANCH, "sha": sha}, timeout=20)
        ok = pr.status_code in (200, 201)
        if ok:
            _GH_CACHE.pop(path, None)
        return ok
    except Exception:
        return False


def gh_write(path, content, message):
    """Записать файл в репозиторий бота (PUT contents с актуальным sha)."""
    if _storage_db(path):
        try:
            _db.kv_set(path, content)
            _GH_CACHE[path] = (content, _time.time())
            return True
        except Exception:
            return False
    if not (GH_TOKEN and GH_REPO):
        return False
    url = f"https://api.github.com/repos/{GH_REPO}/contents/{path}"
    h = {"Authorization": f"token {GH_TOKEN}", "Accept": "application/vnd.github+json"}
    sha = None
    try:
        r = requests.get(f"{url}?ref={GH_BRANCH}", headers=h, timeout=15)
        if r.status_code == 200:
            sha = r.json().get("sha")
    except Exception:
        pass
    body = {"message": message, "branch": GH_BRANCH,
            "content": base64.b64encode(content.encode("utf-8")).decode("ascii")}
    if sha:
        body["sha"] = sha
    try:
        pr = requests.put(url, headers=h, json=body, timeout=20)
        ok = pr.status_code in (200, 201)
        if ok:
            _GH_CACHE[path] = (content, _time.time())   # сразу кладём свежее в кэш
        return ok
    except Exception:
        return False


def _clean_task(s):
    s = s.lstrip("- ")
    for e in ("🔴", "🟡", "🟢", "✅"):
        s = s.replace(e, "")
    s = re.sub(r"\[[^\]]*\]", "", s)
    s = re.sub(r"\(до:\d{4}-\d{2}-\d{2}\)", "", s)
    s = re.sub(r"_\([^)]*\)_", "", s)
    s = re.sub(r"\(T-[^)]*\)", "", s)
    return re.sub(r"\s+", " ", s).strip()


def by_id(uid):
    u = USERS_BY_ID.get(uid)
    if u:
        return {**u, "sections": ROLE_SECTIONS[u["role"]]}
    return {"id": uid, "role": "staff", "companies": [], "sections": ROLE_SECTIONS["staff"]}


def can_company(profile, company):
    if profile["role"] == "owner" or profile.get("companies") == "*":
        return True
    if company in ("Группа", "*", ""):
        return False
    return any(c in company for c in profile.get("companies", []))


# ---------- модель состояния бота → форма фронта ----------
def _parse_section(md, header, done):
    out = []
    if header not in (md or ""):
        return out
    block = md.split(header, 1)[1].split("\n##", 1)[0]
    i = 0
    for ln in block.splitlines():
        ln = ln.strip()
        if not ln.startswith("- "):
            continue
        pr = next((e for e in ("🔴", "🟡", "🟢") if e in ln), "🟡")
        m = re.search(r"\[([^\]]+)\]", ln)
        dm = re.search(r"\(до:(\d{4}-\d{2}-\d{2})\)", ln)
        text = ln.lstrip("- ")
        for e in ("🔴", "🟡", "🟢", "✅"):
            text = text.replace(e, "")
        text = re.sub(r"\[[^\]]*\]", "", text)
        text = re.sub(r"\(до:\d{4}-\d{2}-\d{2}\)", "", text)
        text = re.sub(r"_\([^)]*\)_", "", text)
        text = re.sub(r"\(T-[^)]*\)", "", text).strip()
        i += 1
        out.append({"id": ("d" if done else "t") + str(i), "text": text, "priority": pr,
                    "company": m.group(1) if m else "", "due": dm.group(1) if dm else "", "done": done})
    return out


def parse_tasks(md):
    # активные + последние выполненные (для вкладки «Выполнено»)
    return _parse_section(md, "## Активные", False) + _parse_section(md, "## Выполнено", True)[:40]


def days_until(date_str):
    try:
        return (dt.date.fromisoformat(date_str) - dt.date.today()).days
    except Exception:
        return None


class Login(BaseModel):
    pin: str


def _client_ip(request):
    xff = request.headers.get("x-forwarded-for", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _audit_login(ip, ok, uid=""):
    try:
        ym = dt.date.today().strftime("%Y-%m")
        path = f"state/audit/logins-{ym}.json"
        log = load_json(path, [])
        if not isinstance(log, list):
            log = []
        log.append({"ts": dt.datetime.now(dt.timezone(dt.timedelta(hours=3))).isoformat(timespec="seconds"),
                    "ip": ip, "ok": bool(ok), "uid": uid})
        gh_write(path, json.dumps(log[-500:], ensure_ascii=False, indent=2), "audit: вход")
    except Exception:
        pass


def _do_login(b, request, response):
    ip = _client_ip(request)
    if not _login_gate(ip):
        _audit_login(ip, False, "blocked")
        raise HTTPException(429, "Слишком много попыток. Подождите пару минут.")
    prof = verify_pin(b.pin)
    if not prof:
        _login_fail(ip)
        _audit_login(ip, False)
        raise HTTPException(401, "Неверный PIN")
    _login_ok(ip)
    _audit_login(ip, True, prof["id"])
    _set_session_cookie(response, prof["id"])
    return {"ok": True, "profile": {**prof, "sections": ROLE_SECTIONS[prof["role"]]}}


@app.get("/api/audit/logins")
def audit_logins(user=Depends(require_owner), ym: str = ""):
    ym = ym or dt.date.today().strftime("%Y-%m")
    log = load_json(f"state/audit/logins-{ym}.json", [])
    return {"ok": True, "ym": ym, "logins": (log or [])[-100:][::-1]}


# ---------- Postgres: схема при старте + статус + флаг доменов ----------
def _tasks_db_on():
    return bool(_db and _db.db_available() and os.environ.get("TASKS_BACKEND", "github") == "db")


if _db and _db.db_available():
    try:
        _db.init_schema()
        print("DB: схема инициализирована")
    except Exception as _e:
        print("DB init error:", _e)
    if _tasks_db_on():
        try:
            _db.sync_users(list(USERS_BY_ID.values()))
            if _db.tasks_count() == 0:
                parsed = parse_tasks(gh_read("state/tasks.md") or "")
                n = _db.import_tasks(parsed)
                print(f"DB: задачи импортированы из tasks.md — {n}")
        except Exception as _e:
            print("DB tasks init error:", _e)
    if os.environ.get("STORAGE_BACKEND", "github") == "db":
        # Все домены (кроме данных бота state/workload/*) хранятся в Postgres.
        # Перенос из GitHub — ленивый: каждый файл переносится при первом обращении.
        try:
            print(f"DB: хранилище=Postgres KV (перенос ленивый), в KV записей: {_db.kv_count()}")
        except Exception as _e:
            print("DB storage init error:", _e)


def _cstate_on():
    return bool(_db and _db.db_available() and os.environ.get("CLIENT_STATE_BACKEND", "local") == "db")


@app.get("/api/db/status")
def db_status(user=Depends(require_owner)):
    if not _db:
        return {"connected": False, "reason": "модуль db не загружен"}
    st = _db.status()
    st["tasks_backend"] = "db" if _tasks_db_on() else "github"
    st["client_state_backend"] = "db" if _cstate_on() else "local"
    st["storage_backend"] = "db" if _storage_db("state/_probe") else "github"
    try:
        st["kv_count"] = _db.kv_count() if _db.db_available() else 0
    except Exception:
        st["kv_count"] = 0
    return st


class CState(BaseModel):
    key: str
    data: dict = {}


@app.get("/api/state")
def state_get(key: str = "", user=Depends(current_user)):
    """Клиентское состояние (health|tracker). backend=local → клиент работает на localStorage."""
    if key not in ("health", "tracker"):
        return {"ok": False, "error": "bad key"}
    if not _cstate_on():
        return {"ok": True, "backend": "local", "data": None}
    return {"ok": True, "backend": "db", "data": _db.cstate_get(_prof(user)["id"], key)}


@app.post("/api/state")
def state_set(b: CState, user=Depends(current_user)):
    if b.key not in ("health", "tracker"):
        return {"ok": False, "error": "bad key"}
    if not _cstate_on():
        return {"ok": False, "backend": "local"}
    _db.cstate_set(_prof(user)["id"], b.key, json.dumps(b.data, ensure_ascii=False))
    return {"ok": True, "backend": "db"}


@app.post("/api/auth/login")
def auth_login(b: Login, request: Request, response: Response):
    return _do_login(b, request, response)


@app.post("/api/login")
def login_legacy(b: Login, request: Request, response: Response):
    return _do_login(b, request, response)


@app.get("/api/auth/me")
def auth_me(request: Request, response: Response):
    data = read_session(_token_from(request))
    if not data:
        return {"ok": False}
    prof = by_id(data["uid"])
    _set_session_cookie(response, prof["id"])   # скользящее продление сессии
    return {"ok": True, "profile": {**prof, "sections": ROLE_SECTIONS[prof.get("role", "staff")]}}


@app.post("/api/auth/logout")
def auth_logout(response: Response):
    _clear_session_cookie(response)
    return {"ok": True}


def _occurs(a, d):
    """Является ли дата d (YYYY-MM-DD) вхождением блока сетки a (с учётом повтора)."""
    rep = a.get("repeat")
    if not rep:
        return a.get("date") == d
    if d in (a.get("skip_dates") or []):
        return False
    dates = rep.get("dates")
    if dates:
        return d in dates
    try:
        da = dt.date.fromisoformat(a.get("date", ""))
        dd = dt.date.fromisoformat(d)
    except Exception:
        return False
    if dd < da:
        return False
    every = int(rep.get("every") or 1) or 1
    unit = rep.get("unit") or "day"
    if unit == "day":
        return (dd - da).days % every == 0
    if unit == "week":
        return (dd - da).days % (7 * every) == 0
    if unit == "month":
        if dd.day != da.day:
            return False
        months = (dd.year - da.year) * 12 + (dd.month - da.month)
        return months >= 0 and months % every == 0
    return False


def _occ_done(a, d):
    """Отмечено ли конкретное вхождение (дата d) как выполненное."""
    if a.get("repeat"):
        return d in (a.get("done_dates") or [])
    return bool(a.get("done"))


def _repeat_label(rep):
    if not rep:
        return ""
    if rep.get("dates"):
        return "по датам (%d)" % len(rep["dates"])
    every = int(rep.get("every") or 1)
    unit = {"day": "дн.", "week": "нед.", "month": "мес."}.get(rep.get("unit") or "day", "дн.")
    if unit == "дн." and every == 1:
        return "каждый день"
    return "каждые %d %s" % (every, unit)


def _match_block(a, kind_date, start, text):
    """Найти блок по (старт, текст) с учётом того, что kind_date — дата вхождения."""
    return a.get("start") == start and a.get("text", "") == text and _occurs(a, kind_date)


@app.get("/api/today")
def today(user=Depends(current_user)):
    profile = _prof(user)
    today_iso = dt.date.today().isoformat()
    agenda = []
    for a in load_json("state/agenda.json", []):
        if a.get("start") and _occurs(a, today_iso) and not _occ_done(a, today_iso):
            if profile["role"] == "owner" or can_company(profile, a.get("company", "")) or not a.get("company"):
                agenda.append({"time": a["start"], "text": a.get("text", ""), "icon": "ti-clock"})
    agenda.sort(key=lambda x: x["time"])
    dls = []
    for d in load_json("state/deadlines.json", []):
        if d.get("done"):
            continue
        n = days_until(d.get("date", ""))
        if n is None:
            continue
        if profile["role"] != "owner" and not (can_company(profile, d.get("company", "")) or d.get("company") == "Группа" and profile["role"] == "head"):
            continue
        lvl = "red" if n <= 7 else ("amber" if n <= 21 else "green")
        dls.append({"days": max(n, 0), "text": d.get("text", "")[:22], "level": lvl})
    dls.sort(key=lambda x: x["days"])
    return {"agenda": agenda, "deadlines": dls[:3]}


@app.get("/api/tasks")
def tasks(user=Depends(current_user)):
    profile = _prof(user)
    if _tasks_db_on():
        items = [{"id": str(r["id"]), "text": r["title"], "company": r["company"] or "",
                  "priority": r["priority"], "due": r["due"] or "", "done": r["status"] == "done",
                  "assignee": r["user_id"] or ""} for r in _db.tasks_list()]
    else:
        items = parse_tasks(gh_read("state/tasks.md") or "")
    if profile["role"] == "owner":
        return items
    if profile["role"] == "head":
        return [t for t in items if can_company(profile, t.get("company", ""))]
    return [t for t in items if t.get("assignee") == profile["id"]]


class Done(BaseModel):
    id: str = ""
    text: str = ""


@app.post("/api/tasks/done")
def task_done(b: Done):
    """Закрыть задачу. В режиме БД — по стабильному ID; на GitHub — по тексту (legacy)."""
    if _tasks_db_on() and b.id.isdigit():
        try:
            _db.task_done(b.id)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)[:160]}
    md = gh_read("state/tasks.md") or ""
    target = (b.text or "").strip()
    if not target or "## Активные" not in md:
        return {"ok": False}
    out, removed, in_active = [], None, False
    for l in md.splitlines():
        st = l.strip()
        if st.startswith("## "):
            in_active = (st == "## Активные")
        if in_active and removed is None and st.startswith("- ") and \
                (target in _clean_task(l) or _clean_task(l) in target):
            removed = l
            continue
        out.append(l)
    if removed is None:
        return {"ok": False, "reason": "not found"}
    md2 = "\n".join(out)
    done = f"- ✅ {target} _(закрыто {dt.date.today().isoformat()})_"
    if "## Выполнено" in md2:
        md2 = md2.replace("## Выполнено", "## Выполнено\n" + done, 1)
    else:
        md2 = md2.rstrip() + "\n\n## Выполнено\n" + done + "\n"
    return {"ok": gh_write("state/tasks.md", md2, "app: закрыта задача — " + target[:40])}


class EditTask(BaseModel):
    id: str = ""
    old_text: str = ""
    new_text: str
    due: str = "__keep__"   # "__keep__" — не трогать срок; "" — убрать; "YYYY-MM-DD" — поставить


@app.post("/api/tasks/edit")
def task_edit(b: EditTask):
    if _tasks_db_on() and b.id.isdigit():
        try:
            nw = (b.new_text or "").strip()
            if not nw:
                return {"ok": False}
            _db.task_edit(b.id, nw, b.due)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)[:160]}
    md = gh_read("state/tasks.md") or ""
    target = (b.old_text or "").strip()
    new = (b.new_text or "").strip()
    if not target or not new or "## Активные" not in md:
        return {"ok": False}
    out, edited, in_active = [], False, False
    for l in md.splitlines():
        st = l.strip()
        if st.startswith("## "):
            in_active = (st == "## Активные")
        if in_active and not edited and st.startswith("- ") and (target in _clean_task(l) or _clean_task(l) in target):
            pr = next((e for e in ("🔴", "🟡", "🟢") if e in l), "🟡")
            mc = re.search(r"\[([^\]]+)\]", l)
            mt = re.search(r"_\([^)]*\)_", l)
            md_old = re.search(r"\(до:(\d{4}-\d{2}-\d{2})\)", l)
            comp = mc.group(1) if mc else "Личное"
            tail = (" " + mt.group(0)) if mt else ""
            due_val = (md_old.group(1) if md_old else "") if b.due == "__keep__" else b.due
            due_tag = f" (до:{due_val})" if due_val else ""
            out.append(f"- {pr} [{comp}] {new}{due_tag}{tail}")
            edited = True
            continue
        out.append(l)
    if not edited:
        return {"ok": False, "reason": "not found"}
    return {"ok": gh_write("state/tasks.md", "\n".join(out), "app: правка задачи")}


class Touch(BaseModel):
    name: str


@app.post("/api/deals/touch")
def deal_touch(b: Touch):
    """Отметить касание сделки: last_touch = сегодня. Пишет в deals.json."""
    deals = load_json("state/deals.json", [])
    nm = (b.name or "").strip()
    hit = False
    for d in deals:
        if nm and nm in d.get("name", ""):
            d["last_touch"] = dt.date.today().isoformat()
            hit = True
            break
    if not hit:
        return {"ok": False}
    return {"ok": gh_write("state/deals.json", json.dumps(deals, ensure_ascii=False, indent=2), "app: касание сделки")}


@app.get("/api/deals")
def deals(user=Depends(current_user)):
    profile = _prof(user)
    raw = load_json("state/deals.json", [])
    today_d = dt.date.today()
    out = []
    for d in raw:
        try:
            silent = (today_d - dt.date.fromisoformat(d.get("last_touch", ""))).days
        except Exception:
            silent = 0
        out.append({"name": d.get("name", ""), "company": d.get("company", ""),
                    "stage": d.get("stage", "лид"), "step": d.get("next_step", ""), "silent": silent})
    if profile["role"] == "owner":
        return out
    if profile["role"] == "head":
        return [d for d in out if can_company(profile, d.get("company", ""))]
    return []


@app.get("/api/finance")
def finance(user=Depends(current_user)):
    profile = _prof(user)
    fin = load_json("state/finance.json", {})
    comps = fin.get("companies", {}) or {}
    def num(v):
        return isinstance(v, (int, float))
    if profile["role"] == "owner":
        income = sum((c.get("profit") or 0) * (c.get("share", 1.0) or 1.0)
                     for c in comps.values() if num(c.get("profit")))
        lev = load_json("state/levers.json", {})
        levers = lev.get("levers", []) if isinstance(lev, dict) else []
        return {"scope": "group", "data": {
            "ownerIncome": round(income), "goal": fin.get("goal_income", 5000000),
            "debt": fin.get("debt_total", 0),
            "companies": [{"name": k, "profit": v.get("profit") or 0} for k, v in comps.items() if num(v.get("profit"))],
            "levers": [{"name": l.get("name", ""), "impact": l.get("impact", 0), "progress": l.get("progress", 0)} for l in levers]
        }}
    if profile["role"] == "head":
        cos = [{"name": k, "profit": v.get("profit") or 0} for k, v in comps.items()
               if any(p in k for p in profile.get("companies", [])) and num(v.get("profit"))]
        return {"scope": "company", "data": {"companies": cos}}
    return {"scope": "none", "data": None}


class AddTask(BaseModel):
    text: str
    company: str = ""
    priority: str = "🟡"
    due: str = ""


@app.post("/api/tasks/add")
def task_add(b: AddTask):
    if _tasks_db_on():
        try:
            _db.task_add(b.text.strip(), b.company, b.priority, b.due, _owner_id())
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)[:160]}
    md = gh_read("state/tasks.md") or "# Задачи\n\n## Активные\n\n## Выполнено\n"
    today = dt.date.today().isoformat()
    due_tag = f" (до:{b.due})" if b.due else ""
    line = f"- {b.priority} [{b.company or 'Личное'}] {b.text.strip()}{due_tag} _(добавлено {today})_"
    md2 = md.replace("## Активные", "## Активные\n" + line, 1) if "## Активные" in md \
        else md.rstrip() + "\n\n## Активные\n" + line + "\n"
    return {"ok": gh_write("state/tasks.md", md2, "app: новая задача")}


def _norm_simple(line):
    s = line.lstrip("- ")
    for e in ("🔴", "🟡", "🟢", "✅"):
        s = s.replace(e, "")
    s = re.sub(r"\[[^\]]*\]", " ", s)
    s = re.sub(r"_\([^)]*\)_", " ", s)
    s = re.sub(r"\(T-[^)]*\)", " ", s)
    s = re.sub(r"[^\wа-яё0-9 ]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def _similar(a, b):
    if not a or not b:
        return False
    if a in b or b in a:
        return True
    sa = {w[:5] for w in a.split() if len(w) >= 4}
    sb = {w[:5] for w in b.split() if len(w) >= 4}
    if not sa or not sb:
        return False
    return len(sa & sb) / len(sa | sb) >= 0.6


@app.post("/api/tasks/dedup")
def tasks_dedup():
    if _tasks_db_on():
        return {"ok": True, "removed": 0}   # в БД дубли не копятся (стабильные ID)
    md = gh_read("state/tasks.md") or ""
    if "## Активные" not in md:
        return {"ok": False, "removed": 0}
    head, rest = md.split("## Активные", 1)
    active_block = rest.split("\n##", 1)[0]
    after = rest[len(active_block):]
    lines = [l for l in active_block.splitlines() if l.strip().startswith("- ")]
    norms = [_norm_simple(l) for l in lines]
    n = len(lines)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    for i in range(n):
        for j in range(i + 1, n):
            if _similar(norms[i], norms[j]):
                parent[find(i)] = find(j)
    groups = {}
    for i in range(n):
        groups.setdefault(find(i), []).append(i)
    drop = set()
    for idxs in groups.values():
        if len(idxs) < 2:
            continue
        keep = max(idxs, key=lambda k: len(lines[k]))
        for k in idxs:
            if k != keep:
                drop.add(k)
    if not drop:
        return {"ok": True, "removed": 0}
    kept = [lines[k] for k in range(n) if k not in drop]
    new_md = head + "## Активные\n" + "\n".join(kept) + "\n" + after
    return {"ok": gh_write("state/tasks.md", new_md, f"app: dedup -{len(drop)}"), "removed": len(drop)}


class AddReminder(BaseModel):
    date: str
    time: str = "09:00"
    text: str


@app.post("/api/reminders/add")
def reminder_add(b: AddReminder):
    rems = load_json("state/reminders.json", [])
    rems.append({"when": f"{b.date}T{b.time}", "text": b.text.strip(), "done": False})
    return {"ok": gh_write("state/reminders.json", json.dumps(rems, ensure_ascii=False, indent=2), "app: напоминание")}


class AddBlock(BaseModel):
    date: str
    start: str
    end: str = ""
    text: str
    repeat: dict | None = None      # {every:N, unit:"day|week|month"} | {dates:[...]}


@app.post("/api/agenda/add")
def agenda_add(b: AddBlock):
    ag = load_json("state/agenda.json", [])
    entry = {"date": b.date, "start": b.start, "end": (b.end or None), "text": b.text.strip()}
    if b.repeat and (b.repeat.get("dates") or b.repeat.get("every")):
        entry["repeat"] = b.repeat
        entry["done_dates"] = []
        entry["skip_dates"] = []
    ag.append(entry)
    return {"ok": gh_write("state/agenda.json", json.dumps(ag, ensure_ascii=False, indent=2), "app: блок в сетку")}


@app.get("/api/month")
def month(ym: str = ""):
    ym = ym or dt.date.today().isoformat()[:7]
    tot, pend = {}, {}

    def bump(dte, done):
        tot[dte] = tot.get(dte, 0) + 1
        if not done:
            pend[dte] = pend.get(dte, 0) + 1

    try:
        y0, m0 = int(ym[:4]), int(ym[5:7])
        first = dt.date(y0, m0, 1)
        nxt = dt.date(y0 + 1, 1, 1) if m0 == 12 else dt.date(y0, m0 + 1, 1)
        month_days = [(first + dt.timedelta(days=i)).isoformat() for i in range((nxt - first).days)]
    except Exception:
        month_days = []
    for a in load_json("state/agenda.json", []):
        if not a.get("start"):
            continue
        if a.get("repeat"):
            for dstr in month_days:
                if _occurs(a, dstr):
                    bump(dstr, _occ_done(a, dstr))
        elif str(a.get("date", ""))[:7] == ym:
            bump(a["date"], bool(a.get("done")))
    for r in load_json("state/reminders.json", []):
        w = r.get("when", "")
        if len(w) >= 10 and "T" in w and w[:7] == ym:
            bump(w[:10], bool(r.get("done")))
    status = {d: ("pending" if pend.get(d, 0) > 0 else "done") for d in tot}
    return {"dates": sorted(tot.keys()), "status": status}


class Note(BaseModel):
    text: str


@app.post("/api/note/add")
def note_add(b: Note):
    j = gh_read("state/journal.md") or "# Дневник\n"
    stamp = dt.datetime.now(dt.timezone(dt.timedelta(hours=3))).strftime("%Y-%m-%d %H:%M")
    j = j.rstrip() + f"\n- {stamp} · {b.text.strip()}\n"
    return {"ok": gh_write("state/journal.md", j, "app: заметка")}


class ItemMove(BaseModel):
    kind: str            # "block" | "rem"
    date: str
    start: str
    text: str
    new_date: str
    new_start: str
    new_end: str = ""


@app.post("/api/item/move")
def item_move(b: ItemMove):
    if b.kind == "block":
        ag = load_json("state/agenda.json", [])
        for a in ag:
            if _match_block(a, b.date, b.start, b.text):
                if a.get("repeat"):
                    # перенос одного вхождения серии: снять этот день + разовый блок на новом месте
                    a.setdefault("skip_dates", []).append(b.date)
                    ag.append({"date": b.new_date, "start": b.new_start, "end": (b.new_end or None), "text": b.text})
                else:
                    a["date"] = b.new_date
                    a["start"] = b.new_start
                    a["end"] = b.new_end or None
                return {"ok": gh_write("state/agenda.json", json.dumps(ag, ensure_ascii=False, indent=2), "app: перенос блока")}
        return {"ok": False, "reason": "not found"}
    else:
        rems = load_json("state/reminders.json", [])
        for r in rems:
            w = r.get("when", "")
            if w[:10] == b.date and w[11:16] == b.start and r.get("text", "") == b.text:
                r["when"] = f"{b.new_date}T{b.new_start}"
                return {"ok": gh_write("state/reminders.json", json.dumps(rems, ensure_ascii=False, indent=2), "app: перенос напоминания")}
        return {"ok": False, "reason": "not found"}


class ItemDel(BaseModel):
    kind: str
    date: str
    start: str
    text: str
    scope: str = "one"      # "one" — только этот день; "series" — вся серия повтора


@app.post("/api/item/delete")
def item_delete(b: ItemDel):
    if b.kind == "block":
        ag = load_json("state/agenda.json", [])
        for a in ag:
            if _match_block(a, b.date, b.start, b.text):
                if a.get("repeat") and b.scope != "series":
                    a.setdefault("skip_dates", []).append(b.date)  # убрать только это вхождение
                    return {"ok": gh_write("state/agenda.json", json.dumps(ag, ensure_ascii=False, indent=2), "app: снят день повтора")}
                new = [x for x in ag if x is not a]
                return {"ok": gh_write("state/agenda.json", json.dumps(new, ensure_ascii=False, indent=2), "app: удалён блок")}
        return {"ok": False, "reason": "not found"}
    else:
        rems = load_json("state/reminders.json", [])
        new = [r for r in rems if not (r.get("when", "")[:10] == b.date and r.get("when", "")[11:16] == b.start and r.get("text", "") == b.text)]
        if len(new) == len(rems):
            return {"ok": False, "reason": "not found"}
        return {"ok": gh_write("state/reminders.json", json.dumps(new, ensure_ascii=False, indent=2), "app: удалено напоминание")}


@app.get("/api/journal")
def journal(limit: int = 50):
    j = gh_read("state/journal.md") or ""
    entries = []
    for ln in j.splitlines():
        m = re.match(r"^- (.+?) · (.+)$", ln.strip())
        if m:
            entries.append({"ts": m.group(1), "text": m.group(2)})
    entries.reverse()
    return {"entries": entries[:limit]}


@app.get("/api/day")
def day(user=Depends(current_user), date: str = ""):
    d = date or dt.date.today().isoformat()
    items = []
    for a in load_json("state/agenda.json", []):
        if a.get("start") and _occurs(a, d) and not _occ_done(a, d):
            row = {"start": a["start"], "end": a.get("end"), "text": a.get("text", ""), "kind": "block"}
            if a.get("repeat"):
                row["recurring"] = True
                row["repeat"] = a["repeat"]
                row["repeat_label"] = _repeat_label(a["repeat"])
            items.append(row)
    for r in load_json("state/reminders.json", []):
        if r.get("done"):
            continue
        w = r.get("when", "")
        if len(w) >= 16 and w[:10] == d and "T" in w:
            items.append({"start": w[11:16], "end": None, "text": r.get("text", ""), "kind": "rem"})
    items.sort(key=lambda x: x["start"])

    def m(s):
        try:
            hh, mm = s.split(":")
            return int(hh) * 60 + int(mm)
        except Exception:
            return None
    ws, we = 8 * 60, 22 * 60
    busy = []
    for it in items:
        s = m(it["start"])
        e = m(it["end"]) if it["end"] else (s + 60 if s is not None else None)
        if s is None:
            continue
        busy.append((max(s, ws), min(max(e, s + 15), we)))
    busy = [x for x in busy if x[0] < we and x[1] > ws]
    busy.sort()
    free, cur = [], ws
    for s, e in busy:
        if s > cur:
            free.append(f"{cur//60:02d}:{cur%60:02d}–{s//60:02d}:{s%60:02d}")
        cur = max(cur, e)
    if cur < we:
        free.append(f"{cur//60:02d}:{cur%60:02d}–22:00")
    # выполненные за этот день (из сетки и напоминаний) — для разбора «что сделано»
    done_items = []
    for a in load_json("state/agenda.json", []):
        if a.get("start") and _occurs(a, d) and _occ_done(a, d):
            done_items.append({"start": a["start"], "text": a.get("text", ""), "kind": "block",
                               "recurring": bool(a.get("repeat"))})
    for r in load_json("state/reminders.json", []):
        w = r.get("when", "")
        if r.get("done") and len(w) >= 16 and w[:10] == d and "T" in w:
            done_items.append({"start": w[11:16], "text": r.get("text", ""), "kind": "rem"})
    done_items.sort(key=lambda x: x["start"])
    return {"date": d, "items": items, "free": free, "done": done_items}


@app.get("/api/push/key")
def push_key():
    return {"key": VAPID_PUBLIC, "ready": push_ready()}


class Sub(BaseModel):
    subscription: dict


@app.post("/api/push/subscribe")
def push_subscribe(b: Sub):
    subs = load_json("state/push_subs.json", [])
    ep = (b.subscription or {}).get("endpoint")
    if not ep:
        return {"ok": False}
    if not any(s.get("endpoint") == ep for s in subs):
        subs.append(b.subscription)
        gh_write("state/push_subs.json", json.dumps(subs, ensure_ascii=False, indent=2), "app: push subscribe")
    return {"ok": True, "count": len(subs)}


@app.post("/api/push/test")
def push_test():
    if not push_ready():
        return {"ok": False, "reason": "VAPID_PRIVATE не задан"}
    subs = load_json("state/push_subs.json", [])
    sent = 0
    for s in subs:
        try:
            send_push(s, {"title": "Клиники Столицы", "body": "Тест уведомления — всё работает ✓", "url": "/"})
            sent += 1
        except Exception:
            pass
    return {"ok": True, "sent": sent}


# ---------- Центр уведомлений: настройки каналов ----------
NOTIF_DEFAULTS = {
    "master":        True,
    "morning":       {"on": True, "time": "08:00"},
    "evening":       {"on": True, "time": "21:00"},
    "deadlines":     {"on": True, "time": "09:00", "days": 3},
    "tasks_overdue": {"on": True, "time": "09:00"},
    "agenda_day":    {"on": True, "time": "08:00"},
    "health":        {"on": True, "time": "10:00", "days": 7},
    "aggregator":    {"on": True, "time": "12:00"},
}


def notif_settings():
    s = load_json("state/notif_settings.json", {})
    out = json.loads(json.dumps(NOTIF_DEFAULTS))
    if isinstance(s, dict):
        if "master" in s:
            out["master"] = bool(s["master"])
        for k, v in s.items():
            if k in out and isinstance(v, dict) and isinstance(out.get(k), dict):
                out[k].update(v)
    return out


@app.get("/api/notif/settings")
def notif_get():
    return {"ok": True, "settings": notif_settings()}


class NotifSet(BaseModel):
    settings: dict


@app.post("/api/notif/settings")
def notif_save(b: NotifSet):
    ok = gh_write("state/notif_settings.json", json.dumps(b.settings or {}, ensure_ascii=False, indent=2), "app: настройки уведомлений")
    return {"ok": ok, "settings": notif_settings()}


def _notif_now():
    return dt.datetime.now(dt.timezone(dt.timedelta(hours=3))).replace(tzinfo=None)


def _time_reached(hhmm, now):
    try:
        h, m = [int(x) for x in str(hhmm).split(":")]
    except Exception:
        h, m = 9, 0
    return (now.hour, now.minute) >= (h, m)


def _notif_once_path(ch, d):
    return f"state/notif/sent-{ch}-{d}.json"


def _notif_should(ch, cfg, now):
    if not cfg.get("on"):
        return False
    if not _time_reached(cfg.get("time", "09:00"), now):
        return False
    return load_json(_notif_once_path(ch, now.date().isoformat()), None) is None


def _notif_mark(ch, now, info=""):
    gh_write(_notif_once_path(ch, now.date().isoformat()),
             json.dumps({"ts": now.isoformat(timespec="minutes"), "info": info}, ensure_ascii=False),
             f"notif {ch}")


def _push_all(title, body, url="/"):
    n = 0
    for s in load_json("state/push_subs.json", []):
        try:
            send_push(s, {"title": title, "body": (body or "")[:180], "url": url})
            n += 1
        except Exception:
            pass
    return n


# ---------- сторож бота: пуш владельцу, если бот замолчал (Этап D) ----------
_BOT_WD = {"down": False}


def _bot_watchdog():
    """Читает heartbeat бота (state/heartbeat.json, живёт в GitHub) и шлёт пуш,
       если отметка старше BOT_DOWN_MIN минут. Один пуш на простой + один при восстановлении."""
    if not push_ready():
        return
    hb = load_json("state/heartbeat.json", {})
    try:
        ts = int(hb.get("ts", 0)) if isinstance(hb, dict) else 0
    except Exception:
        ts = 0
    if not ts:
        return                              # отметки ещё нет (бот не обновлён) — не тревожим
    silent = _time.time() - ts
    thr = int(os.environ.get("BOT_DOWN_MIN", "15")) * 60
    if silent > thr and not _BOT_WD["down"]:
        _push_all("⚠️ Бот не отвечает", f"Молчит {int(silent // 60)} мин. Проверьте сервис бота в Railway.", "/")
        _BOT_WD["down"] = True
    elif silent <= thr and _BOT_WD["down"]:
        _push_all("✅ Бот снова в строю", "Бот отвечает штатно.", "/")
        _BOT_WD["down"] = False


class HCheckups(BaseModel):
    items: list = []


@app.post("/api/health/checkups")
def health_checkups_save(b: HCheckups):
    """Клиент синкает график чекапов (только название + дата) для пуш-напоминаний.
       Медицинские значения остаются на устройстве."""
    clean = []
    for it in (b.items or []):
        try:
            t = str(it.get("title") or "").strip()
            d = str(it.get("nextDate") or "").strip()
            if t and d:
                clean.append({"title": t[:60], "nextDate": d})
        except Exception:
            continue
    ok = gh_write("state/health/checkups.json", json.dumps(clean, ensure_ascii=False, indent=2), "health: график чекапов")
    return {"ok": ok, "count": len(clean)}


def _push_morning_loop():
    last = {"d": ""}
    while True:
        try:
            now = dt.datetime.now(dt.timezone.utc)
            today = dt.date.today().isoformat()
            if now.hour == PUSH_HOUR_UTC and last["d"] != today and push_ready():
                subs = load_json("state/push_subs.json", [])
                if subs:
                    near = []
                    for d in load_json("state/deadlines.json", []):
                        if d.get("done"):
                            continue
                        n = days_until(d.get("date", ""))
                        if n is not None and 0 <= n <= 7:
                            near.append(f"{d.get('text', '')[:28]} — {n} дн")
                    body = ("🔴 Горящее: " + "; ".join(near[:3])) if near else "Доброе утро! Открой план на день."
                    for s in subs:
                        try:
                            send_push(s, {"title": "Доброе утро, Иван", "body": body, "url": "/"})
                        except Exception:
                            pass
                last["d"] = today
        except Exception:
            pass
        _time.sleep(180)


threading.Thread(target=_push_morning_loop, daemon=True).start()


# ---------- иконки приложения (генерируются при старте, без бинарей в git) ----------
HERE = os.path.dirname(os.path.abspath(__file__))
WEB = os.path.join(HERE, "webapp")


def ensure_icons():
    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception:
        return
    icodir = os.path.join(WEB, "icons")
    os.makedirs(icodir, exist_ok=True)
    sizes = {"apple-touch-icon.png": 180, "apple-touch-icon-167.png": 167,
             "apple-touch-icon-152.png": 152, "icon-192.png": 192,
             "icon-512.png": 512, "icon-maskable.png": 512}
    for fn, sz in sizes.items():
        p = os.path.join(icodir, fn)
        if os.path.exists(p):
            continue
        img = Image.new("RGB", (sz, sz), (225, 25, 28))   # фирменный красный
        d = ImageDraw.Draw(img)
        try:
            f = ImageFont.load_default(size=int(sz * 0.46))
        except TypeError:
            f = ImageFont.load_default()
        bb = d.textbbox((0, 0), "KC", font=f)
        w, h = bb[2] - bb[0], bb[3] - bb[1]
        d.text((sz / 2 - w / 2 - bb[0], sz / 2 - h / 2 - bb[1]), "KC", font=f, fill=(255, 255, 255))
        img.save(p, "PNG")


ensure_icons()

# ---------- отметить пункт сетки/напоминание выполненным ----------
class ItemDone(BaseModel):
    kind: str
    date: str
    start: str
    text: str


@app.post("/api/item/done")
def item_done(b: ItemDone):
    if b.kind == "block":
        ag = load_json("state/agenda.json", [])
        for a in ag:
            if _match_block(a, b.date, b.start, b.text):
                if a.get("repeat"):
                    dd = a.setdefault("done_dates", [])
                    if b.date not in dd:
                        dd.append(b.date)
                else:
                    a["done"] = True
                return {"ok": gh_write("state/agenda.json", json.dumps(ag, ensure_ascii=False, indent=2), "app: блок выполнен")}
        return {"ok": False, "reason": "not found"}
    else:
        rems = load_json("state/reminders.json", [])
        for r in rems:
            w = r.get("when", "")
            if w[:10] == b.date and w[11:16] == b.start and r.get("text", "") == b.text:
                r["done"] = True
                return {"ok": gh_write("state/reminders.json", json.dumps(rems, ensure_ascii=False, indent=2), "app: напоминание выполнено")}
        return {"ok": False, "reason": "not found"}


@app.post("/api/item/undone")
def item_undone(b: ItemDone):
    """Вернуть выполненный элемент сетки обратно в работу."""
    if b.kind == "block":
        ag = load_json("state/agenda.json", [])
        for a in ag:
            if _match_block(a, b.date, b.start, b.text):
                if a.get("repeat"):
                    a["done_dates"] = [x for x in (a.get("done_dates") or []) if x != b.date]
                else:
                    a["done"] = False
                return {"ok": gh_write("state/agenda.json", json.dumps(ag, ensure_ascii=False, indent=2), "app: блок возвращён")}
        return {"ok": False, "reason": "not found"}
    else:
        rems = load_json("state/reminders.json", [])
        for r in rems:
            w = r.get("when", "")
            if w[:10] == b.date and w[11:16] == b.start and r.get("text", "") == b.text:
                r["done"] = False
                return {"ok": gh_write("state/reminders.json", json.dumps(rems, ensure_ascii=False, indent=2), "app: напоминание возвращено")}
        return {"ok": False, "reason": "not found"}


class ItemEdit(BaseModel):
    kind: str
    date: str
    start: str
    text: str
    new_text: str
    new_start: str
    new_end: str = ""


@app.post("/api/item/edit")
def item_edit(b: ItemEdit):
    """Правка блока на месте (текст/время), сохраняя повтор и историю выполнений."""
    if b.kind == "block":
        ag = load_json("state/agenda.json", [])
        for a in ag:
            if _match_block(a, b.date, b.start, b.text):
                a["start"] = b.new_start
                a["end"] = (b.new_end or None)
                a["text"] = (b.new_text or "").strip() or a.get("text", "")
                return {"ok": gh_write("state/agenda.json", json.dumps(ag, ensure_ascii=False, indent=2), "app: правка блока")}
        return {"ok": False, "reason": "not found"}
    else:
        rems = load_json("state/reminders.json", [])
        for r in rems:
            w = r.get("when", "")
            if w[:10] == b.date and w[11:16] == b.start and r.get("text", "") == b.text:
                r["when"] = f"{b.date}T{b.new_start}"
                r["text"] = (b.new_text or "").strip() or r.get("text", "")
                return {"ok": gh_write("state/reminders.json", json.dumps(rems, ensure_ascii=False, indent=2), "app: правка напоминания")}
        return {"ok": False, "reason": "not found"}


# ---------- дедлайны: список / правка / выполнено / добавить ----------
@app.get("/api/deadlines")
def deadlines_list(user=Depends(current_user)):
    profile = _prof(user)
    raw = load_json("state/deadlines.json", [])
    out = []
    for i, d in enumerate(raw):
        if d.get("done"):
            continue
        if profile["role"] != "owner" and not (can_company(profile, d.get("company", "")) or (d.get("company") == "Группа" and profile["role"] == "head")):
            continue
        n = days_until(d.get("date", ""))
        out.append({"i": i, "date": d.get("date", ""), "text": d.get("text", ""),
                    "company": d.get("company", ""), "days": n})
    out.sort(key=lambda x: (x["days"] is None, x["days"] if x["days"] is not None else 0))
    return out


class DlEdit(BaseModel):
    i: int
    date: str = ""
    text: str = ""


@app.post("/api/deadlines/edit")
def deadline_edit(b: DlEdit):
    raw = load_json("state/deadlines.json", [])
    if b.i < 0 or b.i >= len(raw):
        return {"ok": False, "reason": "bad index"}
    if b.date:
        raw[b.i]["date"] = b.date
    if b.text:
        raw[b.i]["text"] = b.text.strip()
    return {"ok": gh_write("state/deadlines.json", json.dumps(raw, ensure_ascii=False, indent=2), "app: дедлайн изменён")}


class DlDone(BaseModel):
    i: int


@app.post("/api/deadlines/done")
def deadline_done(b: DlDone):
    raw = load_json("state/deadlines.json", [])
    if b.i < 0 or b.i >= len(raw):
        return {"ok": False}
    raw[b.i]["done"] = True
    return {"ok": gh_write("state/deadlines.json", json.dumps(raw, ensure_ascii=False, indent=2), "app: дедлайн закрыт")}


class DlAdd(BaseModel):
    date: str
    text: str
    company: str = ""


@app.post("/api/deadlines/add")
def deadline_add(b: DlAdd):
    raw = load_json("state/deadlines.json", [])
    raw.append({"date": b.date, "text": b.text.strip(), "company": b.company, "done": False})
    return {"ok": gh_write("state/deadlines.json", json.dumps(raw, ensure_ascii=False, indent=2), "app: новый дедлайн")}


# ---------- цели и рычаги ----------
@app.get("/api/goals")
def goals(user=Depends(current_user)):
    bosses = load_json("state/bosses.json", [])
    gl = []
    for g in bosses:
        tot = g.get("total") or 0
        left = g.get("left", tot)
        pct = 0 if not tot else max(0, min(100, round((tot - left) / tot * 100)))
        gl.append({"name": g.get("name", ""), "total": tot, "left": left, "unit": g.get("unit", ""), "pct": pct})
    lev = load_json("state/levers.json", {})
    levers = lev.get("levers", []) if isinstance(lev, dict) else []
    levers = [{"name": l.get("name", ""), "impact": l.get("impact", 0),
               "progress": l.get("progress", 0), "note": l.get("note", "")} for l in levers]
    return {"goals": gl, "levers": levers}


# ---------- план недели ----------
@app.get("/api/weekplan")
def weekplan(user=Depends(current_user)):
    ag = load_json("state/agenda.json", [])
    rems = load_json("state/reminders.json", [])
    today = dt.date.today()
    days = []
    for off in range(7):
        iso = (today + dt.timedelta(days=off)).isoformat()
        items = []
        for a in ag:
            if a.get("date") == iso and a.get("start") and not a.get("done"):
                items.append({"time": a["start"], "text": a.get("text", ""), "kind": "block"})
        for r in rems:
            w = r.get("when", "")
            if not r.get("done") and w[:10] == iso and "T" in w:
                items.append({"time": w[11:16], "text": r.get("text", ""), "kind": "rem"})
        items.sort(key=lambda x: x["time"])
        days.append({"date": iso, "items": items})
    tasks = parse_tasks(gh_read("state/tasks.md") or "")
    tdy = today.isoformat()
    # задачи со сроком — кладём под свой день
    for d in days:
        for t in tasks:
            if not t["done"] and t.get("due") == d["date"]:
                d["items"].append({"time": "", "text": t["text"], "kind": "task"})
        d["items"].sort(key=lambda x: (x["time"] == "", x["time"]))
    # просроченные (дата в прошлом, не выполнено) — предложим перенести
    overdue = []
    for t in tasks:
        if not t["done"] and t.get("due") and t["due"] < tdy:
            overdue.append({"kind": "task", "text": t["text"], "date": t["due"]})
    for a in ag:
        if a.get("start") and not a.get("done") and a.get("date") and a["date"] < tdy:
            overdue.append({"kind": "block", "date": a["date"], "start": a["start"], "text": a.get("text", "")})
    for r in rems:
        w = r.get("when", "")
        if not r.get("done") and "T" in w and len(w) >= 16 and w[:10] < tdy:
            overdue.append({"kind": "rem", "date": w[:10], "start": w[11:16], "text": r.get("text", "")})
    overdue.sort(key=lambda x: x.get("date", ""))
    return {"days": days, "overdue": overdue}


# ---------- привычки и шаги ----------
def _streak(date_strs):
    days = set()
    for s in date_strs:
        try:
            days.add(dt.date.fromisoformat(s))
        except Exception:
            pass
    today = dt.date.today()
    if today in days:
        cur = today
    elif (today - dt.timedelta(days=1)) in days:
        cur = today - dt.timedelta(days=1)
    else:
        return 0
    n = 0
    while cur in days:
        n += 1
        cur -= dt.timedelta(days=1)
    return n


def _ladder_goal(plan, streak):
    ladder = plan.get("ladder") or []
    if not ladder:
        return plan.get("goal", "")
    cur = ladder[0].get("goal", "")
    for lvl in sorted(ladder, key=lambda x: x.get("min", 0)):
        if streak >= lvl.get("min", 0):
            cur = lvl.get("goal", cur)
    return cur


@app.get("/api/habits")
def habits(user=Depends(current_user)):
    plans = load_json("state/habit_plans.json", {})
    log = load_json("state/habits.json", {})
    today = dt.date.today()
    out = []
    for name in sorted(set(list(plans.keys()) + list(log.keys()))):
        dates = log.get(name, [])
        ds = set()
        for s in dates:
            try:
                ds.add(dt.date.fromisoformat(s))
            except Exception:
                pass
        chain = [(today - dt.timedelta(days=i)) in ds for i in range(6, -1, -1)]
        plan = plans.get(name, {})
        st = _streak(dates)
        out.append({"name": name, "streak": st, "chain": chain, "week": sum(chain),
                    "goal": _ladder_goal(plan, st), "anchor": plan.get("anchor", ""),
                    "category": plan.get("category", ""), "done_today": today.isoformat() in {s for s in dates}})
    return {"habits": out}


class HabitDone(BaseModel):
    habit: str


@app.post("/api/habits/done")
def habit_done(b: HabitDone):
    log = load_json("state/habits.json", {})
    today = dt.date.today().isoformat()
    arr = log.get(b.habit, [])
    if today not in arr:
        arr.append(today)
    log[b.habit] = arr
    return {"ok": gh_write("state/habits.json", json.dumps(log, ensure_ascii=False, indent=2), "app: привычка отмечена")}


# ---------- большие цели (год/месяц) ----------
@app.get("/api/biggoals")
def biggoals(user=Depends(current_user)):
    g = load_json("state/big_goals.json", [])
    order = {"open": 0, "done": 1, "miss": 2}
    return sorted(g, key=lambda x: (order.get(x.get("status", "open"), 0),
                                    0 if x.get("scope") == "year" else 1,
                                    x.get("period", "")), reverse=False)


class BGAdd(BaseModel):
    scope: str = "year"
    period: str = ""
    text: str = ""


@app.post("/api/biggoals/add")
def biggoal_add(b: BGAdd):
    import uuid
    g = load_json("state/big_goals.json", [])
    g.append({"id": uuid.uuid4().hex[:8], "scope": b.scope, "period": b.period,
              "text": b.text.strip(), "status": "open", "ts": dt.date.today().isoformat()})
    return {"ok": gh_write("state/big_goals.json", json.dumps(g, ensure_ascii=False, indent=2), "app: большая цель")}


class BGStatus(BaseModel):
    id: str
    status: str


@app.post("/api/biggoals/status")
def biggoal_status(b: BGStatus):
    g = load_json("state/big_goals.json", [])
    hit = False
    for x in g:
        if x.get("id") == b.id:
            x["status"] = b.status
            hit = True
            break
    if not hit:
        return {"ok": False}
    return {"ok": gh_write("state/big_goals.json", json.dumps(g, ensure_ascii=False, indent=2), "app: статус цели")}


class BGDel(BaseModel):
    id: str


@app.post("/api/biggoals/delete")
def biggoal_delete(b: BGDel):
    g = [x for x in load_json("state/big_goals.json", []) if x.get("id") != b.id]
    return {"ok": gh_write("state/big_goals.json", json.dumps(g, ensure_ascii=False, indent=2), "app: цель удалена")}


# ---------- группа компаний: помесячная прибыль по направлениям ----------
DEFAULT_DIRS = {"ПВЛ": 1.0, "Агрегатор Москва": 1.0, "Клиника на Павелецкой": 1.0, "Астрахань": 1.0, "Калмыкия": 0.5}


def _prev_ym(ym):
    try:
        y, m = int(ym[:4]), int(ym[5:7])
    except Exception:
        return ""
    m -= 1
    if m < 1:
        m = 12
        y -= 1
    return f"{y}-{m:02d}"


def _fin_net(v):
    """Чистая по направлению: dict{income,expense}→разница; число→как есть (legacy); иначе None."""
    if isinstance(v, dict):
        try:
            return float(v.get("income") or 0) - float(v.get("expense") or 0)
        except Exception:
            return None
    if isinstance(v, (int, float)):
        return v
    return None


def _agg_month_by_region(ym):
    """Месячная маржа агрегатора (валовая) по регионам из agg-daily."""
    store = load_json(_agg_path(ym), {})
    reg = {}
    for _d, payload in (store or {}).items():
        for r in payload.get("rows", []):
            m = (r.get("revenue", 0) or 0) - (r.get("cost", 0) or 0)
            reg[r.get("region", "—")] = reg.get(r.get("region", "—"), 0) + m
    out = {k: round(v) for k, v in reg.items()}
    out["_total"] = round(sum(reg.values()))
    return out


@app.get("/api/group")
def group(user=Depends(require_owner), ym: str = ""):
    months = load_json("state/finance_months.json", {})
    dirs = load_json("state/finance_dirs.json", {}) or dict(DEFAULT_DIRS)
    allym = sorted(months.keys())
    if not ym:
        ym = allym[-1] if allym else dt.date.today().isoformat()[:7]
    data = months.get(ym, {})
    prev = _prev_ym(ym)
    pdata = months.get(prev, {})
    names = list(dict.fromkeys(list(dirs.keys()) + list(data.keys())))
    agg = _agg_month_by_region(ym)
    rows = []
    total = owner = 0.0
    for n in names:
        v = data.get(n)
        share = dirs.get(n, 1.0)
        net = _fin_net(v)
        inc = v.get("income") if isinstance(v, dict) else (v if isinstance(v, (int, float)) else None)
        exp = v.get("expense") if isinstance(v, dict) else None
        sug = None
        for reg, mv in agg.items():
            if reg != "_total" and reg.lower() in n.lower():
                sug = mv
                break
        rows.append({"name": n, "income": inc, "expense": exp, "net": net,
                     "prev_net": _fin_net(pdata.get(n)), "share": share, "agg_suggest": sug})
        if isinstance(net, (int, float)):
            total += net
            owner += net * share
    trend = [{"ym": y, "total": round(sum((_fin_net(v) or 0) for v in months[y].values()))} for y in allym[-6:]]
    goal = load_json("state/finance.json", {}).get("goal_income", 5000000)
    return {"ym": ym, "prev_ym": prev, "rows": rows, "total": round(total),
            "owner_income": round(owner), "goal": goal, "agg_total": agg.get("_total", 0),
            "months": allym, "trend": trend}


class GroupSave(BaseModel):
    ym: str
    rows: list


@app.post("/api/group/save")
def group_save(b: GroupSave):
    months = load_json("state/finance_months.json", {})
    dirs = load_json("state/finance_dirs.json", {}) or dict(DEFAULT_DIRS)
    md = {}
    for r in b.rows:
        nm = (r.get("name") or "").strip()
        if not nm:
            continue
        inc, exp = r.get("income"), r.get("expense")
        if inc not in (None, "") or exp not in (None, ""):
            def _f(x):
                try:
                    return float(x) if x not in (None, "") else 0.0
                except Exception:
                    return 0.0
            md[nm] = {"income": _f(inc), "expense": _f(exp)}
        else:
            pr = r.get("profit")
            try:
                pr = float(pr) if pr not in (None, "") else None
            except Exception:
                pr = None
            if pr is not None:
                md[nm] = pr
        sh = r.get("share")
        try:
            if sh not in (None, ""):
                dirs[nm] = float(sh)
        except Exception:
            pass
    months[b.ym] = md
    ok1 = gh_write("state/finance_months.json", json.dumps(months, ensure_ascii=False, indent=2), f"app: финансы группы {b.ym}")
    ok2 = gh_write("state/finance_dirs.json", json.dumps(dirs, ensure_ascii=False, indent=2), "app: доли направлений")
    return {"ok": ok1 and ok2}


class ReportUpload(BaseModel):
    kind: str                 # "agg" | "finance"
    filename: str = ""
    data_b64: str
    date: str = ""            # agg: дата данных (по умолчанию вчера)
    ym: str = ""              # finance: месяц


RU_MONTHS = {"январ": 1, "феврал": 2, "март": 3, "апрел": 4, "мая": 5, "май": 5, "июн": 6,
             "июл": 7, "август": 8, "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12}


def _fin_num(v):
    """Число из ячейки: '1 022 448,00 ₽' → 1022448.0. None, если не число."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v or "").strip().replace(" ", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if s in ("", "-", ".", "--"):
        return None
    try:
        return float(s)
    except Exception:
        return None


def _fin_guess_ym(grid, filename="", fallback=""):
    """Месяц отчёта из имени файла или шапки: 'пвл июнь' → 2026-06.
       Числовые ячейки игнорируем, иначе год ловится из сумм (208191.88 → 2081)."""
    fn = (filename or "").lower()
    head = []
    for row in grid[:3]:
        for c in (row or []):
            s = str(c or "").strip()
            if s and _fin_num(s) is None:      # только текстовые ячейки
                head.append(s.lower())
    hay = fn + " " + " ".join(head)
    mon = None
    for k, v in RU_MONTHS.items():
        if k in hay:
            mon = v
            break
    if not mon:
        return fallback or ""
    yr = None
    m = re.search(r"(?<!\d)(20\d{2})(?!\d)", hay)          # явный год: 2026
    if m:
        yr = int(m.group(1))
    else:
        m2 = re.search(r"(?<!\d)(\d{2})(?!\d)", fn)        # 'май 26' — только из имени файла
        if m2 and 20 <= int(m2.group(1)) <= 40:
            yr = 2000 + int(m2.group(1))
    if not yr:
        yr = dt.date.today().year
    return f"{yr}-{mon:02d}"


def _fin_parse_simple(grid, filename=""):
    """Разбор простого отчёта «статья | сумма» со строками ИТОГО (доход) / ИТОГО (расход).
       Работает без ИИ — надёжно для месячных бюджетов направлений."""
    pairs = []
    for row in grid:
        if not row:
            continue
        label = str(row[0] or "").strip()
        val = None
        for c in list(row)[1:5]:
            n = _fin_num(c)
            if n is not None:
                val = n
                break
        if label or val is not None:
            pairs.append((label, val))
    totals = [v for l, v in pairs if l.lower().startswith("итого") and v is not None]
    profit = [v for l, v in pairs if l.lower().startswith("прибыл") and v is not None]
    if len(totals) >= 2:
        income, expense = totals[0], totals[1]
    elif len(totals) == 1 and profit:
        income, expense = totals[0], totals[0] - profit[0]
    else:
        return None
    fn = (filename or "").lower()
    name = "Направление"
    if "пвл" in fn or "павелец" in fn:
        name = "ПВЛ"
    elif "агрегат" in fn:
        name = "Агрегатор Москва"
    elif "астрахан" in fn:
        name = "Астрахань"
    elif "калмык" in fn:
        name = "Калмыкия"
    return {"name": name, "income": round(income, 2), "expense": round(expense, 2),
            "profit": round(profit[0], 2) if profit else round(income - expense, 2)}


def _json_loads_loose(txt):
    """Терпимый разбор JSON от ИИ: снимает markdown и чинит обрыв ответа."""
    t = (txt or "").replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(t)
    except Exception:
        pass
    i = t.rfind("}")
    if i > 0:
        head = t[:i + 1]
        for suf in ("]}", "}]}", "}", ""):
            try:
                return json.loads(head + suf)
            except Exception:
                continue
    raise ValueError("ответ ИИ оборван")


@app.post("/api/report/upload")
def report_upload(b: ReportUpload, user=Depends(require_owner)):
    """Ручная загрузка Excel-отчёта с пометкой типа: агрегатор → в свод; доход/расход → ИИ → на подтверждение."""
    try:
        raw = base64.b64decode(b.data_b64.split(",")[-1])
    except Exception:
        return {"ok": False, "error": "битый файл"}
    if len(raw) > 12 * 1024 * 1024:
        return {"ok": False, "error": "файл больше 12 МБ"}
    try:
        grid = _xls_grid(raw, b.filename)
    except Exception as e:
        return {"ok": False, "error": "не удалось прочитать Excel: " + str(e)[:120]}

    if b.kind == "agg":
        rows = _agg_parse_grid(grid)
        if not rows or sum(r["revenue"] for r in rows) <= 0:
            return {"ok": False, "error": "формат отчёта агрегатора не распознан"}
        d = b.date or (dt.date.today() - dt.timedelta(days=1)).isoformat()
        ym = d[:7]
        store, safe = load_store_safe(_agg_path(ym))
        if not safe:
            return {"ok": False, "error": "чтение свода не удалось — повторите"}
        store[d] = {"rows": rows, "source": "upload:" + (b.filename or ""),
                    "ts": dt.datetime.now().isoformat(timespec="minutes")}
        ok = gh_write(_agg_path(ym), json.dumps(store, ensure_ascii=False, indent=2), f"agg upload {d}")
        rev = sum(r["revenue"] for r in rows)
        cost = sum(r["cost"] for r in rows)
        return {"ok": ok, "kind": "agg", "date": d, "rows": len(rows),
                "revenue": rev, "cost": cost, "margin": rev - cost}

    if b.kind == "finance":
        # 1) Быстрый детерминированный разбор (без ИИ) — для отчётов «статья | сумма» с ИТОГО
        try:
            simple = _fin_parse_simple(grid, b.filename)
        except Exception:
            simple = None
        if simple:
            ym = b.ym or _fin_guess_ym(grid, b.filename) or dt.date.today().strftime("%Y-%m")
            return {"ok": True, "kind": "finance", "ym": ym, "source": "table",
                    "rows": [{"name": simple["name"], "income": simple["income"], "expense": simple["expense"]}],
                    "profit": simple["profit"]}
        # 2) Иначе — ИИ (запасной путь)
        key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not key:
            return {"ok": False, "error": "ANTHROPIC_API_KEY не задан на сервере"}
        text = "\n".join("\t".join(str(c) for c in row) for row in grid[:400])[:12000]
        prompt = ("В таблице — месячный управленческий отчёт по направлениям компании (доходы и расходы). "
                  "Верни СТРОГО JSON без markdown: "
                  '{"ym":"YYYY-MM или пусто","rows":[{"name":"направление","income":число,"expense":число}]}. '
                  "income — совокупный доход направления за месяц, expense — совокупные расходы (рубли, без пробелов). "
                  "Направления бери как в таблице. Чего нет — ставь 0.")
        body = {"model": os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6"), "max_tokens": 4000,
                "messages": [{"role": "user", "content": prompt + "\n\nТАБЛИЦА:\n" + text}]}
        try:
            r = requests.post("https://api.anthropic.com/v1/messages",
                              headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                              json=body, timeout=120)
            if r.status_code != 200:
                return {"ok": False, "error": f"claude {r.status_code}: {r.text[:120]}"}
            txt = "".join(p.get("text", "") for p in r.json().get("content", []) if p.get("type") == "text")
            data = _json_loads_loose(txt)
        except Exception as e:
            return {"ok": False, "error": "разбор ИИ: " + str(e)[:120]}
        ym = b.ym or data.get("ym") or dt.date.today().strftime("%Y-%m")
        parsed = [{"name": (x.get("name") or "").strip(),
                   "income": x.get("income") or 0, "expense": x.get("expense") or 0}
                  for x in (data.get("rows") or []) if (x.get("name") or "").strip()]
        return {"ok": True, "kind": "finance", "ym": ym, "rows": parsed}

    return {"ok": False, "error": "неизвестный тип отчёта"}


# ---------- голос → текст (Whisper) ----------
class STT(BaseModel):
    audio_b64: str
    mime: str = "audio/webm"


@app.post("/api/stt")
def stt(b: STT):
    key = os.environ.get("OPENAI_API_KEY", "")
    if not key:
        return {"ok": False, "error": "OPENAI_API_KEY не задан на сервере"}
    try:
        raw = base64.b64decode(b.audio_b64.split(",")[-1])
    except Exception:
        return {"ok": False, "error": "битый звук"}
    m = (b.mime or "").lower()
    ext = "webm" if "webm" in m else ("m4a" if ("mp4" in m or "m4a" in m or "aac" in m) else ("ogg" if "ogg" in m else "wav"))
    try:
        r = requests.post("https://api.openai.com/v1/audio/transcriptions",
                          headers={"Authorization": f"Bearer {key}"},
                          files={"file": (f"audio.{ext}", raw, b.mime or "audio/webm")},
                          data={"model": "whisper-1", "language": "ru"}, timeout=120)
        if r.status_code == 200:
            return {"ok": True, "text": (r.json().get("text") or "").strip()}
        return {"ok": False, "error": f"whisper {r.status_code}: {r.text[:160]}"}
    except Exception as e:
        return {"ok": False, "error": str(e)[:160]}


# ---------- документы → текст / финансы (Claude Vision) ----------
class Vision(BaseModel):
    image_b64: str
    mime: str = "image/jpeg"
    mode: str = "text"     # "text" — извлечь суть; "finance" — прибыль по направлениям в JSON


@app.post("/api/vision")
def vision(b: Vision):
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        return {"ok": False, "error": "ANTHROPIC_API_KEY не задан на сервере"}
    mime = (b.mime or "").lower()
    if not (mime.startswith("image/") or "pdf" in mime):
        return {"ok": False, "error": "Только фото или PDF"}
    data = b.image_b64.split(",")[-1]
    if len(data) * 3 // 4 > 10 * 1024 * 1024:
        return {"ok": False, "error": "Файл больше 10 МБ"}
    is_pdf = "pdf" in mime
    if b.mode == "finance":
        prompt = ("На изображении/в документе — финансовый отчёт по направлениям компании. "
                  "Верни СТРОГО JSON без пояснений и без markdown: "
                  '{"period":"YYYY-MM или пусто","rows":[{"name":"направление","profit":число_рублей_без_пробелов}]}. '
                  "profit — чистая прибыль за месяц (если есть только выручка и расходы — посчитай прибыль). "
                  "Названия направлений бери как в отчёте.")
    elif b.mode == "labs":
        prompt = ("На фото/в PDF — результаты медицинских анализов. Верни СТРОГО JSON без пояснений и markdown: "
                  '{"date":"YYYY-MM-DD или пусто","tests":[{"marker":"название показателя","value":число,'
                  '"unit":"единицы","ref_min":число_или_null,"ref_max":число_или_null}]}. '
                  "Показатели бери как в бланке (можно по-русски). value — только число (десятичный разделитель — точка). "
                  "Если референс указан диапазоном «a–b» — раздели на ref_min и ref_max; если только одна граница — вторую поставь null. "
                  "Дату исследования возьми из бланка. НЕ придумывай значения, которых нет в документе.")
    else:
        prompt = ("Извлеки суть документа кратко на русском: что это, ключевые суммы, даты, стороны, главное. "
                  "Без воды, до 8 строк.")
    if is_pdf:
        src = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": data}}
    else:
        src = {"type": "image", "source": {"type": "base64", "media_type": b.mime or "image/jpeg", "data": data}}
    body = {"model": os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6"), "max_tokens": 1500,
            "messages": [{"role": "user", "content": [src, {"type": "text", "text": prompt}]}]}
    try:
        r = requests.post("https://api.anthropic.com/v1/messages",
                          headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                          json=body, timeout=120)
        if r.status_code == 200:
            txt = "".join(p.get("text", "") for p in r.json().get("content", []) if p.get("type") == "text")
            return {"ok": True, "text": txt.strip()}
        return {"ok": False, "error": f"claude {r.status_code}: {r.text[:160]}"}
    except Exception as e:
        return {"ok": False, "error": str(e)[:160]}


# ---------- Здоровье: файлы (гибрид) + ИИ-рекомендации ----------
class HFile(BaseModel):
    name: str
    mime: str = "application/octet-stream"
    data_b64: str          # можно с префиксом data:...;base64,


class HFileId(BaseModel):
    id: str


def _hfile_path(fid):
    return f"state/health/files/{fid}.json"


@app.post("/api/health/file")
def health_file_put(b: HFile):
    mime = (b.mime or "").lower()
    if not (mime.startswith("image/") or "pdf" in mime):
        return {"ok": False, "error": "Недопустимый тип файла (только фото или PDF)"}
    if len(b.data_b64.split(",")[-1]) * 3 // 4 > 8 * 1024 * 1024:
        return {"ok": False, "error": "Файл больше 8 МБ"}
    fid = dt.datetime.now().strftime("%Y%m%d%H%M%S") + "-" + str(_time.time_ns() % 100000)
    rec = {"id": fid, "name": b.name, "mime": b.mime,
           "date": dt.date.today().isoformat(), "data": b.data_b64}
    ok = gh_write(_hfile_path(fid), json.dumps(rec, ensure_ascii=False), f"health: файл {b.name[:40]}")
    return {"ok": ok, "id": fid, "name": b.name, "mime": b.mime, "date": rec["date"]}


@app.get("/api/health/file")
def health_file_get(id: str = ""):
    rec = load_json(_hfile_path(id), None)
    if not rec:
        return {"ok": False, "error": "файл не найден"}
    return {"ok": True, "name": rec.get("name"), "mime": rec.get("mime"),
            "date": rec.get("date"), "data": rec.get("data")}


@app.post("/api/health/file/delete")
def health_file_delete(b: HFileId):
    if not b.id:
        return {"ok": False}
    return {"ok": gh_delete(_hfile_path(b.id), "health: удалён файл")}


class HAdvice(BaseModel):
    summary: str = ""


@app.post("/api/health/advice")
def health_advice(b: HAdvice):
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        return {"ok": False, "error": "ANTHROPIC_API_KEY не задан на сервере"}
    sys = ("Ты — заботливый помощник по здоровью и образу жизни. НЕ ставь диагнозы, НЕ назначай "
           "лечение или препараты. Опирайся на общие принципы профилактики и ЗОЖ. Тон спокойный и "
           "поддерживающий, без запугивания. Любое отклонение — повод обсудить с врачом, а не диагноз. "
           "Верни СТРОГО JSON без markdown: "
           '{"overview":"1-2 фразы общего вывода","lifestyle":["совет по образу жизни/питанию/режиму", ...],'
           '"ask_doctor":["что стоит уточнить у врача", ...],"retest":["что и примерно когда имеет смысл пересдать", ...]}. '
           "По-русски, коротко и конкретно, 3-6 пунктов в каждом списке. Если данных мало — скажи это в overview.")
    body = {"model": os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6"), "max_tokens": 1200,
            "system": sys,
            "messages": [{"role": "user", "content": "Показатели и чекапы человека:\n" + (b.summary or "нет данных")}]}
    try:
        r = requests.post("https://api.anthropic.com/v1/messages",
                          headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                          json=body, timeout=120)
        if r.status_code != 200:
            return {"ok": False, "error": f"claude {r.status_code}: {r.text[:160]}"}
        txt = "".join(p.get("text", "") for p in r.json().get("content", []) if p.get("type") == "text")
        try:
            data = json.loads(txt.replace("```json", "").replace("```", "").strip())
        except Exception:
            data = {"overview": txt.strip(), "lifestyle": [], "ask_doctor": [], "retest": []}
        data["ok"] = True
        return data
    except Exception as e:
        return {"ok": False, "error": str(e)[:160]}


# ---------- объединённый главный экран (1 запрос вместо 3) ----------
@app.get("/api/home")
def home_combined(user=Depends(current_user)):
    return {"agenda": today(user)["agenda"], "deadlines": deadlines_list(user), "tasks": tasks(user)}


# ---------- умный разбор: текст/голос → действия (Claude) ----------
class BrainIn(BaseModel):
    text: str


@app.post("/api/brain")
def brain_parse(b: BrainIn):
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        return {"ok": False, "error": "ANTHROPIC_API_KEY не задан на сервере"}
    text = (b.text or "").strip()
    if not text:
        return {"ok": False, "error": "пусто"}
    today_iso = dt.date.today().isoformat()
    prompt = (
        f"Ты ассистент-планировщик. Сегодня {today_iso} (формат YYYY-MM-DD). "
        "Разбери сообщение владельца на конкретные действия и верни СТРОГО JSON без markdown и пояснений: "
        '{"actions":[...]}. Допустимые элементы actions: '
        '{"type":"task","text":"...","due":"YYYY-MM-DD"(опц),"priority":"🔴"|"🟡"|"🟢"(опц)}; '
        '{"type":"reminder","date":"YYYY-MM-DD","time":"HH:MM","text":"..."}; '
        '{"type":"block","date":"YYYY-MM-DD","start":"HH:MM","end":"HH:MM"(опц),"text":"..."}; '
        '{"type":"note","text":"..."}; {"type":"done","match":"часть текста задачи, которую закрыть"}. '
        "Относительные даты («завтра», «в пятницу», «через неделю») переведи в YYYY-MM-DD от сегодня. "
        "Если для напоминания нет времени — 09:00. Если действий нет — пустой список. "
        "Сообщение: " + text)
    body = {"model": os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6"), "max_tokens": 1200,
            "messages": [{"role": "user", "content": prompt}]}
    try:
        r = requests.post("https://api.anthropic.com/v1/messages",
                          headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                          json=body, timeout=60)
        if r.status_code != 200:
            return {"ok": False, "error": f"claude {r.status_code}: {r.text[:140]}"}
        raw = "".join(p.get("text", "") for p in r.json().get("content", []) if p.get("type") == "text")
        raw = raw.replace("```json", "").replace("```", "").strip()
        actions = (json.loads(raw) or {}).get("actions", [])
    except Exception as e:
        return {"ok": False, "error": "разбор: " + str(e)[:140]}
    cnt = {"task": 0, "reminder": 0, "block": 0, "note": 0, "done": 0}
    for a in actions:
        ty = a.get("type")
        try:
            if ty == "task":
                task_add(AddTask(text=a.get("text", ""), due=a.get("due", "") or "", priority=a.get("priority", "🟡") or "🟡"))
            elif ty == "reminder":
                reminder_add(AddReminder(date=a.get("date", ""), time=a.get("time", "09:00") or "09:00", text=a.get("text", "")))
            elif ty == "block":
                agenda_add(AddBlock(date=a.get("date", ""), start=a.get("start", ""), end=a.get("end", "") or "", text=a.get("text", "")))
            elif ty == "note":
                note_add(Note(text=a.get("text", "")))
            elif ty == "done":
                task_done(Done(text=a.get("match", "")))
            else:
                continue
            cnt[ty] = cnt.get(ty, 0) + 1
        except Exception:
            pass
    parts = []
    names = {"task": "задач", "reminder": "напом.", "block": "в сетку", "note": "заметок", "done": "закрыто"}
    for k, v in cnt.items():
        if v:
            parts.append(f"{v} {names[k]}")
    summary = ("Готово: " + ", ".join(parts)) if parts else "Не нашёл, что добавить — уточни формулировку."
    return {"ok": True, "summary": summary, "count": sum(cnt.values())}


# ---------- ПИЛОТ ПВЛ: данные бота → приложение владельца ----------
PVL_POS = {"админ": "Администратор", "врач": "Врач", "медсестра": "Медсестра", "руководитель": "Руководитель"}
PVL_STATUS = {"ok": "🟢 в норме", "over": "🟡 перегружен", "idle": "⚪ простаивает", "issue": "🔴 есть вопросы"}


def _pvl_recent(days):
    days = max(int(days), 1)
    today = dt.date.today()
    months = {(today - dt.timedelta(days=d)).strftime("%Y-%m") for d in range(days)}
    out = []
    for mo in months:
        out += load_json(f"state/workload/checkins-ПВЛ-{mo}.json", [])
    cutoff = (today - dt.timedelta(days=days - 1)).isoformat()
    return [c for c in out if c.get("date", "") >= cutoff]


def _pvl_ints(s):
    return [int(x) for x in re.findall(r"\d+", s or "")]


def _pvl_label(e):
    return "Руководитель" if e.get("role") == "head" else PVL_POS.get(e.get("position"), e.get("position"))


def _pvl_report_data(days):
    emps = load_json("state/workload/employees.json", [])
    team = [{"name": e.get("name"), "position": e.get("position"), "role": e.get("role"),
             "since": e.get("since"), "label": _pvl_label(e)} for e in emps]
    checks = _pvl_recent(days)
    by = {str(e["chat_id"]): {"emp": e, "ev": []} for e in emps if e.get("chat_id")}
    for c in checks:
        cid = str(c.get("chat_id"))
        if cid in by and c.get("type") == "evening":
            by[cid]["ev"].append(c)
    load_rows, quiet = [], []
    for cid, d in by.items():
        e = d["emp"]; ev = d["ev"]
        loads = [x["load"] for x in ev if isinstance(x.get("load"), int)]
        avg = round(sum(loads) / len(loads), 1) if loads else None
        load_rows.append({"name": e.get("name"), "label": _pvl_label(e), "marks": len(ev), "avg": avg})
        if len(ev) < max(1, days // 2):
            quiet.append(e.get("name"))
    vol = {}
    for cid, d in by.items():
        pos = d["emp"].get("position")
        for x in d["ev"]:
            ints = _pvl_ints(x.get("nums"))
            if ints:
                vol.setdefault(pos, [0, 0, 0])
                for i, v in enumerate(ints[:3]):
                    vol[pos][i] += v
    volumes = {}
    if "админ" in vol:
        a = vol["админ"]; volumes["Администраторы"] = f"звонки {a[0]}, записи {a[1]}, дозвоны {a[2]}"
    if "врач" in vol:
        volumes["Врачи"] = f"приёмов {vol['врач'][0]}"
    if "медсестра" in vol:
        volumes["Медсёстры"] = f"процедур {vol['медсестра'][0]}"
    blk, nich, fix = [], [], []
    for cid, d in by.items():
        nm = d["emp"].get("name", "")
        for x in d["ev"]:
            ans = (x.get("qans") or "").strip()
            if not ans or ans in ("-", "—", "нет", "Нет"):
                continue
            row = {"name": nm, "text": ans}
            if x.get("qkey") == "blocker":
                blk.append(row)
            elif x.get("qkey") == "nichye":
                nich.append(row)
            elif x.get("qkey") == "fix":
                fix.append(row)
    ideas = [{"name": c.get("name"), "text": c.get("text")} for c in checks if c.get("type") == "feedback"]
    marks = load_json("state/workload/marks-ПВЛ.json", [])
    last_marks, marks_week = [], ""
    if marks:
        marks_week = marks[-1].get("week", "")
        for m in marks[-1].get("marks", []):
            last_marks.append({"name": m.get("name"), "status": PVL_STATUS.get(m.get("status"), m.get("status")),
                               "note": m.get("note", "")})
    return {"team": team, "load": load_rows, "quiet": quiet, "volumes": volumes,
            "blockers": blk, "nichye": nich, "fix": fix, "ideas": ideas,
            "marks": last_marks, "marks_week": marks_week, "checkins": len(checks)}


def _pvl_ai(data):
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    probs = []
    for r in data["blockers"]:
        probs.append("МЕШАЛО/ЗАВИСЛО: " + r["text"])
    for r in data["nichye"]:
        probs.append("НИЧЬЁ: " + r["text"])
    for r in data["fix"]:
        probs.append("ПОЧИНИТЬ: " + r["text"])
    for r in data["ideas"]:
        probs.append("ИДЕЯ: " + r["text"])
    if not key:
        return {"ok": False, "error": "ANTHROPIC_API_KEY не задан на сервере"}
    if not probs:
        return {"ok": True, "core": [], "tasks": [], "instructions": [],
                "summary": "Пока мало сигналов от команды — как накопятся ответы, появится анализ."}
    prompt = (
        "Ты операционный аналитик клиники. Ниже — реплики сотрудников из ежедневных чек-инов "
        "(что мешало/зависло, что «ничьё», что починить, идеи). Сгруппируй повторяющееся в несколько "
        "КОРНЕВЫХ тем («ядро проблем»), по каждой укажи частоту, и предложи решения. "
        "Верни СТРОГО JSON без markdown и пояснений: "
        '{"core":[{"theme":"короткое название темы","count":число,"detail":"1 фраза сути"}],'
        '"tasks":[{"text":"конкретная задача в повелительном наклонении","priority":"🔴"|"🟡"|"🟢"}],'
        '"instructions":[{"text":"что закрепить регламентом/должностной, чтобы не повторялось"}],'
        '"summary":"2-3 предложения главного вывода для владельца"}. '
        "Будь конкретным и кратким, на русском. Реплики:\n- " + "\n- ".join(probs[:120]))
    body = {"model": os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6"), "max_tokens": 1600,
            "messages": [{"role": "user", "content": prompt}]}
    try:
        r = requests.post("https://api.anthropic.com/v1/messages",
                          headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                          json=body, timeout=90)
        if r.status_code != 200:
            return {"ok": False, "error": f"claude {r.status_code}"}
        raw = "".join(p.get("text", "") for p in r.json().get("content", []) if p.get("type") == "text")
        raw = raw.replace("```json", "").replace("```", "").strip()
        j = json.loads(raw)
        j["ok"] = True
        return j
    except Exception as e:
        return {"ok": False, "error": str(e)[:140]}


@app.get("/api/pvl/team")
def pvl_team(user=Depends(require_owner)):
    return {"ok": True, "team": _pvl_report_data(7)["team"]}


@app.get("/api/pvl/report")
def pvl_report(user=Depends(require_owner), days: int = 7):
    days = 1 if int(days) <= 1 else 7
    data = _pvl_report_data(days)
    data["ai"] = _pvl_ai(data)
    data["ok"] = True
    data["days"] = days
    return data


# ---------- АГРЕГАТОР: дневная выручка/себестоимость/маржа по клиникам ----------
def _agg_path(ym):
    return f"state/finance/agg-daily-{ym}.json"


def _agg_with_margin(rows):
    out = []
    for r in rows:
        rev = float(r.get("revenue") or 0)
        cost = float(r.get("cost") or 0)
        m = rev - cost
        out.append({"region": r.get("region") or "—", "clinic": r.get("clinic") or "—",
                    "revenue": round(rev), "cost": round(cost), "margin": round(m),
                    "pct": round(m / rev * 100, 1) if rev else 0})
    return out


class FinIngest(BaseModel):
    date: str
    rows: list = []
    token: str = ""
    source: str = ""


@app.post("/api/finance/ingest")
def finance_ingest(b: FinIngest):
    need = os.environ.get("INGEST_TOKEN", "")
    if not need:
        # Токен не задан → приём внешних заносов ВЫКЛЮЧЕН (не пропускаем без проверки).
        # Штатный путь — /api/report/upload (под сессией) и почтовый _agg_email_pull.
        return JSONResponse({"ok": False, "error": "Внешний импорт отключён: не задан INGEST_TOKEN"},
                            status_code=503)
    if not hmac.compare_digest(b.token or "", need):
        return JSONResponse({"ok": False, "error": "Неверный токен"}, status_code=403)
    try:
        d = dt.date.fromisoformat(b.date)
    except Exception:
        return {"ok": False, "error": "Дата нужна в формате YYYY-MM-DD"}
    ym = d.strftime("%Y-%m")
    store, safe = load_store_safe(_agg_path(ym))
    if not safe:
        return {"ok": False, "error": "Не удалось прочитать текущий свод из хранилища — занос отменён, чтобы не потерять уже внесённые дни. Повторите позже."}
    clean = []
    for r in (b.rows or []):
        try:
            clean.append({"region": str(r.get("region") or "—"), "clinic": str(r.get("clinic") or "—"),
                          "revenue": round(float(r.get("revenue") or 0)), "cost": round(float(r.get("cost") or 0))})
        except Exception:
            continue
    store[b.date] = {"rows": clean, "source": b.source or "",
                     "ts": dt.datetime.now().isoformat(timespec="minutes")}
    ok = gh_write(_agg_path(ym), json.dumps(store, ensure_ascii=False, indent=2), f"agg ingest {b.date}")
    rev = sum(r["revenue"] for r in clean)
    cost = sum(r["cost"] for r in clean)
    return {"ok": ok, "date": b.date, "clinics": len(clean), "revenue": rev, "cost": cost,
            "margin": rev - cost, "pct": round((rev - cost) / rev * 100, 1) if rev else 0}


def _agg_fixed(ym):
    """Месячные постоянные расходы (для пути к безубыточности). months[ym] или default."""
    cfg = load_json("state/agg_fixed.json", {}) or {}
    if not isinstance(cfg, dict):
        return 0.0
    m = cfg.get("months", {}) or {}
    try:
        return float(m[ym]) if ym in m else float(cfg.get("default", 0) or 0)
    except Exception:
        return 0.0


def _days_in_month(ym):
    try:
        y, mo = int(ym[:4]), int(ym[5:7])
        first = dt.date(y, mo, 1)
        nxt = dt.date(y + 1, 1, 1) if mo == 12 else dt.date(y, mo + 1, 1)
        return (nxt - first).days
    except Exception:
        return 30


class AggFixed(BaseModel):
    amount: float
    ym: str = ""
    scope: str = "default"    # "default" — на все месяцы; "month" — только на ym


@app.post("/api/finance/fixed")
def finance_fixed(b: AggFixed, user=Depends(require_owner)):
    cfg = load_json("state/agg_fixed.json", {}) or {}
    if not isinstance(cfg, dict):
        cfg = {}
    cfg.setdefault("months", {})
    if b.scope == "month" and b.ym:
        cfg["months"][b.ym] = float(b.amount)
    else:
        cfg["default"] = float(b.amount)
    ok = gh_write("state/agg_fixed.json", json.dumps(cfg, ensure_ascii=False, indent=2), "app: постоянные расходы")
    return {"ok": ok}


@app.get("/api/finance/agg")
def finance_agg(user=Depends(require_owner), ym: str = "", mode: str = "month", date: str = ""):
    ym = ym or dt.date.today().strftime("%Y-%m")
    store = load_json(_agg_path(ym), {})
    dates = sorted(store.keys())
    if mode == "day":
        d = date or (dates[-1] if dates else "")
        rows = _agg_with_margin((store.get(d, {}) or {}).get("rows", []))
    else:
        acc = {}
        for _dd, payload in store.items():
            for r in payload.get("rows", []):
                key = (r.get("region", "—"), r.get("clinic", "—"))
                a = acc.setdefault(key, {"region": key[0], "clinic": key[1], "revenue": 0, "cost": 0})
                a["revenue"] += r.get("revenue", 0)
                a["cost"] += r.get("cost", 0)
        rows = _agg_with_margin(list(acc.values()))
        d = ""
    rows.sort(key=lambda x: x["revenue"], reverse=True)
    trev = sum(r["revenue"] for r in rows)
    tcost = sum(r["cost"] for r in rows)
    totals = {"revenue": trev, "cost": tcost, "margin": trev - tcost,
              "pct": round((trev - tcost) / trev * 100, 1) if trev else 0}
    # путь к безубыточности → прибыли (считаем по всему месяцу, независимо от режима)
    m_rev = sum((r.get("revenue", 0) or 0) for p in store.values() for r in p.get("rows", []))
    m_cost = sum((r.get("cost", 0) or 0) for p in store.values() for r in p.get("rows", []))
    m_margin = m_rev - m_cost
    fixed = _agg_fixed(ym)
    ddata, dmonth = len(dates), _days_in_month(ym)
    projected = round(m_margin / ddata * dmonth) if ddata else 0
    breakeven = {"fixed": round(fixed), "month_margin": round(m_margin),
                 "to_breakeven": round(max(0, fixed - m_margin)),
                 "profit": round(m_margin - fixed),
                 "pct": round(min(100, m_margin / fixed * 100), 1) if fixed > 0 else 0,
                 "reached": (m_margin >= fixed) if fixed > 0 else False,
                 "projected": projected,
                 "projected_profit": round(projected - fixed) if fixed > 0 else projected,
                 "days_data": ddata, "days_month": dmonth}
    return {"ok": True, "ym": ym, "mode": mode, "date": d, "dates": dates, "rows": rows,
            "totals": totals, "breakeven": breakeven}


# ---------- АГЕНТ: ежедневный отчёт по марже с почты (IMAP + .xls) ----------
def _agg_num(s):
    s = str(s or "").replace("\n", "").replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
    try:
        return float(s)
    except Exception:
        return None


def _agg_region(cat, name):
    t = (str(name) + " " + str(cat))
    kal = ["Калмык", "Элист", "Лагань", "Городовиков", "Цаган-Аман", "Яшкул", "Яшалт", "Комсомольск", "Приютное"]
    chn = ["Грозн", "Гудермес", "Хасавюрт", "Ачхой", "Серновод", "СклифЛаб", "АРЦ", "НЕОМЕД", "Чечн", "Вайнах"]
    kbr = ["Нальчик", "Баксан", "Прохладн", "Учкекен", "Майский", "КБР", "Головко", "Назрань"]
    msk = ["Павелецк", "Каширск", "Москв", "Профосмотр", "Европейск", "Чайхон", "Титан"]
    for k in kal:
        if k in t:
            return "Калмыкия"
    for k in chn:
        if k in t:
            return "Чечня"
    for k in kbr:
        if k in t:
            return "КБР"
    if "Астрахан" in t:
        return "Астрахань"
    for k in msk:
        if k in t:
            return "Москва"
    return "Прочее"


def _agg_segment(cat):
    return "корпы" if "корп" in str(cat).lower() else "медцентры"


def _xls_grid(raw, filename=""):
    """Единая сетка ячеек (строки×колонки как строки) для .xls (xlrd) и .xlsx (openpyxl)."""
    fn = (filename or "").lower()
    is_xlsx = fn.endswith(".xlsx") or raw[:2] == b"PK"   # .xlsx — это zip (сигнатура PK)
    if is_xlsx:
        import openpyxl
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
        sh = wb[wb.sheetnames[0]]
        grid = []
        for row in sh.iter_rows(values_only=True):
            grid.append(["" if c is None else str(c) for c in row])
        return grid
    import xlrd
    book = xlrd.open_workbook(file_contents=raw)
    sh = book.sheet_by_index(0)
    return [[str(sh.cell_value(ri, ci)) for ci in range(sh.ncols)] for ri in range(sh.nrows)]


def _agg_parse_grid(grid):
    """Разбор сетки отчёта агрегатора -> строки по (регион, сегмент)."""
    def cell(row, ci):
        return row[ci] if ci < len(row) else ""
    hdr = None
    cols = {}
    for ri in range(min(len(grid), 30)):
        vals = [str(c).strip().lower() for c in grid[ri]]
        joined = " ".join(vals)
        if "наименование" in joined and ("счет" in joined or "счёт" in joined):
            hdr = ri
            for ci, v in enumerate(vals):
                if "наименование" in v:
                    cols["name"] = ci
                elif ("по счет" in v or "по счёт" in v or "сумма по счет" in v) and "rev" not in cols:
                    cols["rev"] = ci
                elif "себестоим" in v:
                    cols["cost"] = ci
            break
    if hdr is None or not all(k in cols for k in ("name", "rev", "cost")):
        return []
    agg = {}
    cat = None
    for ri in range(hdr + 1, len(grid)):
        row = grid[ri]
        name = str(cell(row, cols["name"])).replace("\n", " ").strip()
        if not name:
            continue
        if "ИТОГО" in name.upper():
            continue
        rev = _agg_num(cell(row, cols["rev"]))
        cost = _agg_num(cell(row, cols["cost"]))
        if rev is None and cost is None:
            cat = name
            continue
        reg = _agg_region(cat, name)
        seg = _agg_segment(cat)
        a = agg.setdefault((reg, seg), [0, 0])
        a[0] += rev or 0
        a[1] += cost or 0
    return [{"region": r, "clinic": s, "revenue": round(v[0]), "cost": round(v[1])}
            for (r, s), v in agg.items()]


def _agg_parse_xls(raw):
    return _agg_parse_grid(_xls_grid(raw, ""))


def _agg_fetch_candidates(limit=6):
    """Последние письма от отправителя с вложением .xls/.xlsx — новые первыми.
       Возвращает (список [(raw, meta)], ошибка|None)."""
    import imaplib
    import email as _email
    user = os.environ.get("MAIL_USER", "")
    pwd = os.environ.get("MAIL_APP_PASSWORD", "")
    if not (user and pwd):
        return [], "MAIL_USER / MAIL_APP_PASSWORD не заданы"
    sender = os.environ.get("MAIL_FROM", "result@stoclinic.ru")
    host = os.environ.get("MAIL_IMAP", "imap.gmail.com")
    out = []
    try:
        M = imaplib.IMAP4_SSL(host)
        M.login(user, pwd)
        M.select("INBOX")
        typ, dat = M.search(None, f'(FROM "{sender}")')
        ids = dat[0].split()
        if not ids:
            M.logout()
            return [], f"письма от {sender} не найдены"
        for mid in reversed(ids[-limit:]):        # новые первыми
            typ, msgdat = M.fetch(mid, "(RFC822)")
            msg = _email.message_from_bytes(msgdat[0][1])
            try:
                tup = _email.utils.parsedate_to_datetime(msg.get("Date"))
                dd = (tup.date() - dt.timedelta(days=1)).isoformat()
            except Exception:
                dd = (dt.date.today() - dt.timedelta(days=1)).isoformat()
            for part in msg.walk():
                fn = part.get_filename() or ""
                try:
                    import email.header as _eh
                    if fn:
                        fn = str(_eh.make_header(_eh.decode_header(fn)))
                except Exception:
                    pass
                if fn.lower().endswith(".xls") or fn.lower().endswith(".xlsx"):
                    payload = part.get_payload(decode=True)
                    if payload:
                        out.append((payload, {"date": dd, "file": fn, "from": sender}))
                        break
        M.logout()
    except Exception as e:
        if not out:
            return [], "IMAP: " + str(e)[:160]
    return out, (None if out else "вложение .xls не найдено")


def _agg_email_pull():
    cands, err = _agg_fetch_candidates(int(os.environ.get("AGG_SCAN_EMAILS", "4")))
    if not cands:
        return {"ok": False, "error": err or "нет писем с отчётом"}
    # Берём НОВЕЙШИЙ отчёт, который удалось разобрать (маленькие дни/выходные — норм,
    # порогов по сумме нет). Скан нескольких писем — на случай, если у самого свежего
    # письма вложение не читается: тогда берём следующее по свежести.
    chosen = None
    for raw, meta in cands:                       # новые первыми
        try:
            rows = _agg_parse_xls(raw)
        except Exception:
            rows = []
        rev = sum(r["revenue"] for r in rows) if rows else 0
        if rows and rev > 0:
            chosen = (rows, meta, rev)
            break
    if not chosen:
        return {"ok": False, "error": f"не удалось разобрать отчёт из последних {len(cands)} писем"}
    rows, meta, rev = chosen
    d = meta.get("date") or dt.date.today().isoformat()
    ym = d[:7]
    cost = sum(r["cost"] for r in rows)
    store, safe = load_store_safe(_agg_path(ym))
    if not safe:
        return {"ok": False, "error": "Чтение месячного свода не удалось — занос отменён во избежание потери данных. Следующая попытка — в очередной запуск агента."}
    prev = store.get(d) or {}
    prev_rev = sum(r.get("revenue", 0) for r in prev.get("rows", []))
    prev_cost = sum(r.get("cost", 0) for r in prev.get("rows", []))
    base = {"date": d, "rows": len(rows), "revenue": round(rev), "cost": round(cost),
            "margin": round(rev - cost), "pct": round((rev - cost) / rev * 100, 1) if rev else 0,
            "file": meta.get("file", ""), "scanned": len(cands)}
    if prev and abs(prev_rev - rev) < 1 and abs(prev_cost - cost) < 1:
        return {"ok": True, "unchanged": True, **base}   # данные за эту дату уже есть — не переписываем
    store[d] = {"rows": rows, "source": "email:" + meta.get("from", ""),
                "ts": dt.datetime.now().isoformat(timespec="minutes")}
    ok = gh_write(_agg_path(ym), json.dumps(store, ensure_ascii=False, indent=2), f"agg email {d}")
    return {"ok": ok, **base}


@app.get("/api/finance/pull")
def finance_pull(user=Depends(require_owner)):
    return _agg_email_pull()


@app.get("/api/finance/backfill")
def finance_backfill(user=Depends(require_owner), n: int = 45):
    """Перепроверка: сканирует последние n писем от отправителя, разбирает КАЖДОЕ
       и восстанавливает суточные данные по всем найденным датам (пропущенные — добавляет,
       изменившиеся — обновляет, совпадающие — оставляет). Быстрый способ сверить месяц."""
    cands, err = _agg_fetch_candidates(max(1, min(int(n), 90)))
    if not cands:
        return {"ok": False, "error": err or "нет писем с отчётом"}
    seen = set()
    by_month = {}          # ym -> {date: (rows, rev, cost)}
    for raw, meta in cands:                       # новые первыми — на дату берём новейшее письмо
        d = meta.get("date")
        if not d or d in seen:
            continue
        try:
            rows = _agg_parse_xls(raw)
        except Exception:
            rows = []
        rev = sum(r["revenue"] for r in rows) if rows else 0
        if not rows or rev <= 0:
            continue
        seen.add(d)
        by_month.setdefault(d[:7], {})[d] = (rows, rev, sum(r["cost"] for r in rows))
    detail, written = [], 0
    for ym, days in by_month.items():
        store, safe = load_store_safe(_agg_path(ym))
        if not safe:
            detail.append({"month": ym, "status": "чтение не удалось — пропущено"})
            continue
        changed = False
        for d, (rows, rev, cost) in days.items():
            prev = store.get(d) or {}
            prev_rev = sum(r.get("revenue", 0) for r in prev.get("rows", []))
            prev_cost = sum(r.get("cost", 0) for r in prev.get("rows", []))
            if prev and abs(prev_rev - rev) < 1 and abs(prev_cost - cost) < 1:
                detail.append({"date": d, "revenue": round(rev), "status": "совпало"})
                continue
            store[d] = {"rows": rows, "source": "backfill", "ts": dt.datetime.now().isoformat(timespec="minutes")}
            changed, written = True, written + 1
            detail.append({"date": d, "revenue": round(rev), "status": "обновлено" if prev else "добавлено"})
        if changed:
            gh_write(_agg_path(ym), json.dumps(store, ensure_ascii=False, indent=2), f"agg backfill {ym}")
    detail.sort(key=lambda x: x.get("date", ""))
    return {"ok": True, "scanned": len(cands), "dates_found": len(seen), "written": written, "detail": detail}


_agg_attempt = {"ts": 0.0}


def _agg_email_daily():
    if not os.environ.get("MAIL_APP_PASSWORD"):
        return
    now = dt.datetime.now(dt.timezone(dt.timedelta(hours=3))).replace(tzinfo=None)
    if now.hour < 7:
        return
    # тянем ВЕСЬ ДЕНЬ, не чаще раза в 20 мин: как только приходит свежий отчёт —
    # он подхватывается сразу, а не ждёт до завтра. Идемпотентно: если данные за дату
    # уже есть, _agg_email_pull возвращает unchanged и повторно не пишет.
    if _time.time() - _agg_attempt["ts"] < 1200:
        return
    _agg_attempt["ts"] = _time.time()
    res = _agg_email_pull()
    if not res.get("ok"):
        print(f"[{now.strftime('%Y-%m-%d %H:%M')}] agg pull не удался: {res.get('error')}")


# ---------- пуш напоминаний по времени (как у бота, в приложение) ----------
# ---------- ИИ-ПОМОЩНИК: разбор всех процессов владельца ----------
def _owner_id():
    for u in USERS_BY_ID.values():
        if u.get("role") == "owner":
            return u["id"]
    return "ivan"


def _assistant_context(user, health):
    L = []
    try:
        ts = [x for x in tasks(user) if not x.get("done")]
        if ts:
            L.append("ЗАДАЧИ активные (до 20): " + "; ".join((x.get("priority", "") + x.get("text", "")) + (f" [до {x.get('due')}]" if x.get("due") else "") for x in ts[:20]))
    except Exception:
        pass
    try:
        dls = deadlines_list(user)
        if dls:
            L.append("ДЕДЛАЙНЫ: " + "; ".join(f"{d.get('text')} — {d.get('days')}д" for d in dls[:15]))
    except Exception:
        pass
    try:
        dd = deals(user)
        if dd:
            L.append("СДЕЛКИ (воронка): " + "; ".join(f"{x.get('name')} [{x.get('stage')}, тишина {x.get('silent')}д, шаг: {x.get('step', '')}]" for x in dd[:15]))
    except Exception:
        pass
    try:
        f = finance(user)
        if f.get("scope") == "group":
            g = f["data"]
            L.append(f"ФИНАНСЫ: личный доход {g.get('ownerIncome')} из цели {g.get('goal')}, долг {g.get('debt')}. Прибыль направлений: " + ", ".join(f"{c.get('name')}:{c.get('profit')}" for c in g.get("companies", [])))
    except Exception:
        pass
    try:
        ym = dt.date.today().strftime("%Y-%m")
        store = load_json(_agg_path(ym), {})
        if store:
            acc = {}
            for _d, p in store.items():
                for r in p.get("rows", []):
                    a = acc.setdefault(r.get("clinic", "—"), [0, 0])
                    a[0] += r.get("revenue", 0)
                    a[1] += r.get("cost", 0)
            L.append("АГРЕГАТОР за месяц (клиника: выручка/себест/маржа%): " + "; ".join(f"{k}: {v[0]}/{v[1]}/{round((v[0] - v[1]) / v[0] * 100, 1) if v[0] else 0}%" for k, v in acc.items()))
    except Exception:
        pass
    try:
        p = _pvl_report_data(7)
        probs = [x.get("text", "") for x in (p.get("blockers", []) + p.get("nichye", []) + p.get("fix", []))]
        L.append(f"ПВЛ: в боте {len(p.get('team', []))} чел.; сигналы за 7 дн.: " + ("; ".join(probs[:10]) or "нет"))
        if p.get("quiet"):
            L.append("ПВЛ мало отмечаются: " + ", ".join(p["quiet"]))
    except Exception:
        pass
    if health:
        L.append(f"ЗДОРОВЬЕ: индекс контроля {health.get('index', '?')}/100; ближайший чекап {health.get('nextCheckup', '—')}; вне диапазона: {', '.join(health.get('attention', [])) or 'нет'}; просрочено чекапов: {health.get('overdue', 0)}")
        if health.get("habits"):
            L.append("ПРИВЫЧКИ/ТРЕКЕР: " + str(health.get("habits")))
        if health.get("dayReport"):
            L.append("ОТЧЁТ ПО ДНЮ: " + str(health.get("dayReport")))
        if health.get("weight"):
            L.append("ВЕС: " + str(health.get("weight")))
    return "\n".join(L) or "Данных пока мало."


def _assistant_call(key, mode, ctx, question):
    if mode == "ask":
        prompt = ("Ты — личный операционный ассистент владельца группы клиник «Клиники Столицы». "
                  "Ответь кратко и по делу на вопрос, опираясь на данные ниже; если данных не хватает — так и скажи. "
                  "Без медицинских диагнозов и без юридических заключений. Обычный текст, на русском.\n\nВОПРОС: " + question + "\n\nДАННЫЕ:\n" + ctx)
        maxt = 900
    else:
        prompt = ("Ты — личный операционный ассистент владельца группы клиник «Клиники Столицы». "
                  "По данным ниже дай разбор и рекомендации. Верни СТРОГО JSON без markdown: "
                  '{"summary":"2-3 предложения: главное на сегодня и на что смотреть",'
                  '"comments":[{"area":"Задачи|Дедлайны|Сделки|Финансы|Агрегатор|ПВЛ|Здоровье","text":"короткий комментарий и конкретное действие"}],'
                  '"priorities":[{"text":"конкретное действие на сегодня"}]}. '
                  "Будь конкретным, на русском, без воды. По здоровью — без диагнозов, только «обсудить со специалистом».\n\nДАННЫЕ:\n" + ctx)
        maxt = 1600
    body = {"model": os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6"), "max_tokens": maxt,
            "messages": [{"role": "user", "content": prompt}]}
    try:
        r = requests.post("https://api.anthropic.com/v1/messages",
                          headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                          json=body, timeout=90)
        if r.status_code != 200:
            return {"ok": False, "error": f"claude {r.status_code}"}
        raw = "".join(p.get("text", "") for p in r.json().get("content", []) if p.get("type") == "text")
        if mode == "ask":
            return {"ok": True, "answer": raw.strip()}
        raw = raw.replace("```json", "").replace("```", "").strip()
        j = json.loads(raw)
        j["ok"] = True
        j["date"] = dt.date.today().isoformat()
        return j
    except Exception as e:
        return {"ok": False, "error": str(e)[:140]}


class AssistantIn(BaseModel):
    user: str = ""
    mode: str = "cached"
    question: str = ""
    health: dict = {}


@app.post("/api/assistant")
def assistant(b: AssistantIn, user=Depends(require_owner)):
    today_iso = dt.date.today().isoformat()
    cache_path = f"state/assistant/digest-{today_iso}.json"
    if b.mode == "cached":
        c = load_json(cache_path, None)
        return {"ok": True, "cached": bool(c), "digest": c}
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        return {"ok": False, "error": "ANTHROPIC_API_KEY не задан на сервере"}
    ctx = _assistant_context(user["id"], b.health or {})
    if b.mode == "ask":
        return _assistant_call(key, "ask", ctx, b.question or "")
    dg = _assistant_call(key, "digest", ctx, "")
    if dg.get("ok"):
        gh_write(cache_path, json.dumps(dg, ensure_ascii=False, indent=2), f"assistant digest {today_iso}")
    return dg


# ================= КОНТЕНТ-СИСТЕМА v1: агент «Аналитик вирусности» =================
_CONTENT_IDEAS = "state/content/ideas.json"


class ContentIn(BaseModel):
    url: str = ""
    text: str = ""
    note: str = ""


def _content_prompt(src):
    return (
        "Ты — аналитик вирусного видео-контента для группы МЕДИЦИНСКИХ клиник «Клиники Столицы» "
        "(клиники, лаборатории, профосмотры, медкнижки, охрана труда, корпоративная медицина, HR, wellness). "
        "Разбери присланный ролик/пост конкурента или смежной ниши и верни СТРОГО JSON без markdown:\n"
        '{"theme":"тема одной фразой",'
        '"hook":"приём захвата внимания в первые 3 секунды",'
        '"structure":"сценарная структура, 1-2 предложения",'
        '"visual":"ключевой визуальный приём",'
        '"cta":"призыв к действию в оригинале",'
        '"emotion":"базовая эмоция",'
        '"why_viral":"почему зашло, 1-2 предложения",'
        '"applicability":[{"direction":"Клиника|Профосмотры|Медкнижки|B2B|HR/охрана труда","idea":"как перенести МЕХАНИКУ (не копию) на это направление"}],'
        '"compliance":{"art24_ok":true,"flags":["нарушения ст.24 ФЗ О рекламе, если есть: гарантия результата, обещание излечения, запугивание здорового, обращение к несовершеннолетним, недостоверные медутверждения"],'
        '"needs_erid":false,"needs_erid_reason":"почему это реклама и нужна маркировка, либо почему это органика",'
        '"disclaimer":"Имеются противопоказания, необходима консультация специалиста",'
        '"verdict":"можно адаптировать|адаптировать с правками|не подходит бренду"}}\n'
        "Правила: переносим МЕХАНИКУ, не копируем. Для медицины консервативно — любая гарантия/«излечение»/запугивание => art24_ok:false. "
        "Рекламный характер (продвижение конкретной платной услуги) => needs_erid:true; инфо-образовательный контент в своём сообществе обычно органика => needs_erid:false. "
        "Коротко, по-русски.\n\nИСТОЧНИК (ссылка и/или описание ролика):\n" + src)


@app.post("/api/content/analyze")
def content_analyze(b: ContentIn, user=Depends(require_owner)):
    src = "\n".join(x for x in [b.url.strip(), b.note.strip(), b.text.strip()] if x)
    if not src:
        return {"ok": False, "error": "Дайте ссылку или описание ролика"}
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        return {"ok": False, "error": "ANTHROPIC_API_KEY не задан на сервере"}
    body = {"model": os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6"), "max_tokens": 1600,
            "messages": [{"role": "user", "content": _content_prompt(src)}]}
    try:
        r = requests.post("https://api.anthropic.com/v1/messages",
                          headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                          json=body, timeout=90)
        if r.status_code != 200:
            return {"ok": False, "error": f"claude {r.status_code}"}
        raw = "".join(p.get("text", "") for p in r.json().get("content", []) if p.get("type") == "text")
        card = json.loads(raw.replace("```json", "").replace("```", "").strip())
    except Exception as e:
        return {"ok": False, "error": str(e)[:140]}
    card["id"] = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    card["created"] = dt.datetime.now(dt.timezone(dt.timedelta(hours=3))).isoformat(timespec="minutes")
    card["source"] = {"url": b.url.strip(), "note": b.note.strip()}
    card["status"] = "new"
    lst = load_json(_CONTENT_IDEAS, [])
    if not isinstance(lst, list):
        lst = []
    lst.insert(0, card)
    gh_write(_CONTENT_IDEAS, json.dumps(lst[:200], ensure_ascii=False, indent=2), "content: idea")
    return {"ok": True, "card": card}


@app.get("/api/content/ideas")
def content_ideas(user=Depends(require_owner)):
    lst = load_json(_CONTENT_IDEAS, [])
    return {"ok": True, "ideas": lst if isinstance(lst, list) else []}


class ContentDel(BaseModel):
    id: str


@app.post("/api/content/idea/delete")
def content_idea_delete(b: ContentDel, user=Depends(require_owner)):
    lst = load_json(_CONTENT_IDEAS, [])
    lst = [c for c in lst if c.get("id") != b.id] if isinstance(lst, list) else []
    ok = gh_write(_CONTENT_IDEAS, json.dumps(lst, ensure_ascii=False, indent=2), "content: del")
    return {"ok": ok}


def _assistant_daily():
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        return
    now = _notif_now()
    st = notif_settings()
    mcfg = st.get("morning", {})
    if not _time_reached(mcfg.get("time", "08:00"), now):
        return
    cp = f"state/assistant/digest-{now.date().isoformat()}.json"
    if load_json(cp, None) is not None:
        return
    ctx = _assistant_context(_owner_id(), {})
    dg = _assistant_call(key, "digest", ctx, "")
    if dg.get("ok"):
        gh_write(cp, json.dumps(dg, ensure_ascii=False, indent=2), "assistant: daily digest")
        if push_ready() and st.get("master") and mcfg.get("on"):
            _push_all("☀️ Главное на сегодня", dg.get("summary", "") or "", "/")


def _evening_push():
    st = notif_settings()
    ecfg = st.get("evening", {})
    if not (st.get("master") and ecfg.get("on") and push_ready()):
        return
    now = _notif_now()
    if not _time_reached(ecfg.get("time", "21:00"), now):
        return
    mark = f"state/assistant/evening-{now.date().isoformat()}.json"
    if load_json(mark, None) is not None:
        return
    sent = _push_all("🌙 Отчёт по дню", "Отметься: что сделал и не успел + трекер привычек.", "/")
    gh_write(mark, json.dumps({"sent": sent, "ts": now.isoformat(timespec="minutes")}, ensure_ascii=False), "evening push")


def _task_overdue_list():
    tdy = dt.date.today().isoformat()
    return [t for t in tasks(_owner_id()) if not t.get("done") and t.get("due") and str(t["due"]) < tdy]


def _agg_last_date():
    ym = dt.date.today().strftime("%Y-%m")
    dates = sorted(load_json(_agg_path(ym), {}).keys())
    if not dates:
        pm = (dt.date.today().replace(day=1) - dt.timedelta(days=1)).strftime("%Y-%m")
        dates = sorted(load_json(_agg_path(pm), {}).keys())
    return dates[-1] if dates else ""


def _daily_notifications():
    if not push_ready():
        return
    st = notif_settings()
    if not st.get("master"):
        return
    now = _notif_now()
    owner = _owner_id()
    tdy = dt.date.today()

    ch = "deadlines"; cfg = st.get(ch, {})
    if _notif_should(ch, cfg, now):
        days = int(cfg.get("days", 3) or 3)
        soon = [d for d in deadlines_list(owner) if d.get("days") is not None and d["days"] <= days]
        if soon:
            body = "; ".join((d["text"][:40] + " — " + ("сегодня" if d["days"] == 0 else f"через {d['days']}д")) for d in soon[:5])
            _push_all("⏳ Дедлайны", body, "/")
        _notif_mark(ch, now)

    ch = "tasks_overdue"; cfg = st.get(ch, {})
    if _notif_should(ch, cfg, now):
        od = _task_overdue_list()
        if od:
            _push_all("🔴 Просроченные задачи", f"{len(od)}: " + "; ".join((t.get("text", "") or "")[:40] for t in od[:4]), "/")
        _notif_mark(ch, now)

    ch = "agenda_day"; cfg = st.get(ch, {})
    if _notif_should(ch, cfg, now):
        ag = today(owner)["agenda"]
        if ag:
            _push_all("🗓 План на день", "; ".join(f"{a['time']} {a['text'][:32]}" for a in ag[:6]), "/")
        _notif_mark(ch, now)

    ch = "health"; cfg = st.get(ch, {})
    if _notif_should(ch, cfg, now):
        days = int(cfg.get("days", 7) or 7)
        due = []
        for c in load_json("state/health/checkups.json", []):
            try:
                dl = (dt.date.fromisoformat(c["nextDate"]) - tdy).days
            except Exception:
                continue
            if dl <= days:
                due.append(c["title"] + (" — сегодня" if dl == 0 else (f" — просрочено {-dl}д" if dl < 0 else f" — через {dl}д")))
        if due:
            _push_all("🩺 Чек-апы и анализы", "; ".join(due[:5]), "/")
        _notif_mark(ch, now)

    ch = "aggregator"; cfg = st.get(ch, {})
    if _notif_should(ch, cfg, now):
        last = _agg_last_date()
        yday = (tdy - dt.timedelta(days=1)).isoformat()
        if last and last < yday:
            _push_all("📉 Агрегатор", f"Данные за {yday} ещё не пришли (последние — {last}). Проверь отчёт.", "/")
        _notif_mark(ch, now)


def _push_reminders_loop():
    while True:
        try:
            if push_ready():
                rems = load_json("state/reminders.json", [])
                now = dt.datetime.now(dt.timezone(dt.timedelta(hours=3))).replace(tzinfo=None)
                due, changed = [], False
                for r in rems:
                    if r.get("done") or r.get("pushed"):
                        continue
                    w = r.get("when", "")
                    try:
                        when = dt.datetime.fromisoformat(w)
                    except Exception:
                        continue
                    if when <= now:
                        due.append(r)
                        r["pushed"] = True
                        changed = True
                if changed:
                    subs = load_json("state/push_subs.json", [])
                    for r in due:
                        for s in subs:
                            try:
                                send_push(s, {"title": "⏰ Напоминание", "body": r.get("text", ""), "url": "/"})
                            except Exception:
                                pass
                    gh_write("state/reminders.json", json.dumps(rems, ensure_ascii=False, indent=2), "app: напоминания отправлены")
        except Exception as e:
            print(f"push reminders (напоминания): {e}")
        # каждое задание — в своём try, чтобы сбой одного НЕ блокировал остальные
        # (раньше падение _assistant_daily не давало запускаться _agg_email_daily)
        for _job in (_assistant_daily, _evening_push, _daily_notifications, _agg_email_daily, _bot_watchdog):
            try:
                _job()
            except Exception as e:
                print(f"push loop {_job.__name__}: {e}")
        _time.sleep(60)


threading.Thread(target=_push_reminders_loop, daemon=True).start()


# ---------- статика PWA ----------
if os.path.isdir(WEB):
    app.mount("/", StaticFiles(directory=WEB, html=True), name="web")
